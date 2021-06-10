'''
Author: your name
Date: 2021-06-01 19:58:05
LastEditTime: 2021-06-09 18:26:50
Description: file content
██████╗  █████╗ ████████╗ █████╗
██╔══██╗██╔══██╗╚══██╔══╝██╔══██╗
██║  ██║███████║   ██║   ███████║
██║  ██║██╔══██║   ██║   ██╔══██║
██████╔╝██║  ██║   ██║   ██║  ██║
╚═════╝ ╚═╝  ╚═╝   ╚═╝   ╚═╝  ╚═╝
'''
import openpyxl
import linecache
import os
import csv
import math
import copy


import analytic_complexity


# 通过txt获取包含可疑度的excel-start
def get_project_sus(version, project, CODE_FILE_PATH):  # version为in,project为小写
    # 目标存储位置
    TARGET_EXCEL_PATH = '/Users/lvlaxjh/code/CBFL/data/%s/' % (project)
    # 包含真实缺陷
    TRUE_FAULT_FILE_PATH = '/Volumes/Elements/D4j/%s/%s/gzoltars/%s/%s/fault.txt' % (
        project.capitalize(), str(version), project.capitalize(), str(version))
    PROJECT_FILE_PATH = '/Volumes/Elements/D4j/%s/%s/gzoltars/%s/%s/Dstar_%s_%s.txt' % (
        project.capitalize(), str(version), project.capitalize(), str(version), project.capitalize(), str(version))  # 原文件
    # java路径需要修改
    # CODE_FILE_PATH = '/Volumes/Elements/DSatr/%s-direct/%s-%s/src/java/' % (
    #     project, project, str(version))  # java文件父目录
    projectFile = open(PROJECT_FILE_PATH, 'r')  # 项目文件
    projectFileLineList = projectFile.readlines()
    trueFaultFile = open(TRUE_FAULT_FILE_PATH, 'r')  # 真实缺陷文件
    trueFaultFileLineList = trueFaultFile.readlines()
    excel = openpyxl.Workbook()  # 目标存储excel文件
    sheet = excel.worksheets[0]  # 表
    sheetHead = ['project_path', 'version', 'lines',
                 'statement', 'suspicious', 'faulty']
    for i in range(len(sheetHead)):  # 写入表头
        sheet.cell(1, i+1).value = sheetHead[i]
    excelLine = 2  # excel检索行
    for i in projectFileLineList:
        fileLineContentList = i.rstrip("\n").split('#')
        # 去除$-start
        if '$' in fileLineContentList[1]:
            fileLineContentList[1] = fileLineContentList[1][:fileLineContentList[1].find(
                '$')]
            # project_path
            sheet.cell(excelLine, 1).value = fileLineContentList[1]
        else:
            # project_path
            sheet.cell(excelLine, 1).value = fileLineContentList[1]
        # 去除$-end
        sheet.cell(excelLine, 2).value = version  # version
        sheet.cell(excelLine, 3).value = fileLineContentList[2]  # lines
        # 获取代码-start
        try:
            lineOfCode = linecache.getline(
                CODE_FILE_PATH+fileLineContentList[1].replace(".", "/")+'.java', int(fileLineContentList[2]))
        except Exception as e:
            files = open(CODE_FILE_PATH +
                         fileLineContentList[1].replace(".", "/")+'.java', 'r', encoding='iso-8859-1')
            codeFileList = files.readlines()
            lineOfCode = codeFileList[int(fileLineContentList[2])-1]
        # 获取代码-end
        sheet.cell(excelLine, 4).value = lineOfCode  # statement
        sheet.cell(excelLine, 5).value = fileLineContentList[3]  # suspicious
        # 判断该语句是否为真实缺陷-start
        isFault = False  # 是否为真实缺陷，默认为False
        for j in trueFaultFileLineList:
            trueFaultfileLineContentList = j.rstrip("\n").split('#')
            if fileLineContentList[1]+'.java' == trueFaultfileLineContentList[0] and fileLineContentList[2] == trueFaultfileLineContentList[1]:
                isFault = True
            else:
                isFault = False
        if isFault:
            sheet.cell(excelLine, 6).value = 1  # faulty
            print(project+'-'+str(version) +
                  fileLineContentList[1]+'-'+fileLineContentList[2]+' -> is true fault')
        else:
            sheet.cell(excelLine, 6).value = 0  # faulty
        # 判断该语句是否为真实缺陷-end
        excelLine += 1
    excel.save(TARGET_EXCEL_PATH+project+'-'+str(version)+'.xlsx')
    # print('step1 get_project_sus '+project+'-'+str(version)+'.xlsx'+' -> ok')
# 通过txt获取包含可疑度的excel-end


# 运行java文件生成json的AST树-start
def run_java_get_ASTjson(version, project, CODE_FILE_PATH):
    EclipseAST = '/Users/lvlaxjh/code/CBFL/AST/EclipseAST2'
    AST_TO_JSON_APP_JAVA = '/Users/lvlaxjh/code/CBFL/AST/EclipseAST2/App.java'
    EXCEL_PATH = '/Users/lvlaxjh/code/CBFL/data/%s/%s-%s.xlsx' % (
        project, project, str(version))  # 每个版本excel的路径
    JAVA_FATHER_PATH = CODE_FILE_PATH  # java文件父路径
    excel = openpyxl.load_workbook(EXCEL_PATH)  # excel文件
    sheet = excel.worksheets[0]  # 表
    line = 2  # sheet 行
    javaFileList = []
    javaFileNames = []
    # 代码文件路径-start
    while True:
        if sheet.cell(line, 1).value != None:
            filePath = str(sheet.cell(line, 1).value).replace(".", "/")
            if filePath not in javaFileList:
                javaFileNames.append(filePath.split('/')[-1])
                javaFileList.append(filePath)
        else:
            break
        line += 1
    # 代码文件路径-end
    for i in range(len(javaFileList)):
        paths = ''
        for j in javaFileList[i].split('/')[:-1]:
            paths += j
            paths += '/'
        JAVA_FILE_PATH = JAVA_FATHER_PATH+paths
        JAVA_FILE = javaFileNames[i]
        file = open(AST_TO_JSON_APP_JAVA, 'r+')
        # 修改java代码中文件路径-start
        fileList = file.readlines()
        fileList[51] = '		'+'String MAIN_PATH = \"'+JAVA_FILE_PATH+'\";\n'
        fileList[52] = '		'+'String JAVA_FILE = \"'+JAVA_FILE+'\";\n'
        file.close()
        file = open(AST_TO_JSON_APP_JAVA, 'w+')
        file.writelines(fileList)
        file.close()
        # 修改java代码中文件路径-end
        osres = os.popen(  # 运行java
            'cd '+EclipseAST+' && javac -Xlint:deprecation -cp \'lib/*\' *.java && java -Xss4m -cp .:\'lib/*\' App "$@"')
        print(osres.read())
# 运行java文件生成json的AST树-end


# 将可疑度大于1的设置为1-start
def set_sus_one(version, project):
    csvSaveFile = open('/Users/lvlaxjh/code/CBFL/data/%s/csv/sus1/%s%s.csv' % (project, project, str(version)),
                       'w', encoding="utf-8", newline="")
    csvWriter = csv.writer(csvSaveFile)
    csvWriter.writerow(['dataset', 'project', 'path', 'version', 'codeLine', 'statement', 'varTotal', 'optTotal', 'array', 'bracketDepth',
                                   'bracketTotal', 'keywordTotal', 'methodTotal', 'typeTotal', 'logic', 'lengthEle', 'lengthWord', 'depth', 'suspicious', 'accuracy'])

    with open('/Users/lvlaxjh/code/CBFL/data/%s/csv/%s%s.csv' % (project, project, str(version))) as f:
        csvFile = csv.reader(f)
        csvLine = 1  # csv行
        for row in csvFile:
            if csvLine != 1:
                saveLine = row
                if float(row[18]) > 1:  # 大于1，归为1
                    saveLine[18] = 1
                csvWriter.writerow(saveLine)
            csvLine += 1
    csvSaveFile.close()
    print('set_sus_one %s%s.csv -> ok' % (project, str(version)))
# 将可疑度大于1的设置为1-end


# 获取可疑度排名第一并存储-start
def get_sus_top(version, project, topSusList, includeAll):
    # includeAll:是否将全部真实缺陷（无论排名）存储（非必须使用，只针对真实缺陷过少情况）
    with open('/Users/lvlaxjh/code/CBFL/data/%s/csv/sus1/%s%s.csv' % (project, project, str(version))) as f:
        csvFile = csv.reader(f)
        csvLine = 1  # csv行
        sortRow = sorted(list(csvFile)[1:], key=lambda x: float(
            x[18]), reverse=True)  # 排序行
        topSus = float(sortRow[0][18])  # 取可疑度最大
        # 取列表中最大可疑度-start
        for i in sortRow:
            if topSus == float(i[18]):
                topSusList.append(i)
            if includeAll and str(i[19]) == '1' and i not in topSusList:
                topSusList.append(i)
        # 取列表中最大可疑度-end
    print('get_sus_top %s%s.csv -> ok' % (project, str(version)))


def save_top_sus(topSusList, project):
    SAVE_CSV = '/Users/lvlaxjh/code/CBFL/data/%s/csv/top_sus_%s.csv' % (
        project, project)
    csvFile = open(SAVE_CSV, 'w', encoding="utf-8", newline="")
    csvWriter = csv.writer(csvFile)
    csvWriter.writerow(['dataset', 'project', 'path', 'version', 'codeLine', 'statement', 'varTotal', 'optTotal', 'array', 'bracketDepth',
                                   'bracketTotal', 'keywordTotal', 'methodTotal', 'typeTotal', 'logic', 'lengthEle', 'lengthWord', 'depth', 'suspicious', 'accuracy'])
    for i in topSusList:
        csvWriter.writerow(i)
    csvFile.close()
# 获取可疑度排名第一并存储-end


# 切割数据集用于交叉验证-start
def cut_data_for_test_train(project, miniK):  # 当数据量小于10时，设置的k折
    CSV_PATH = '/Users/lvlaxjh/code/CBFL/data/%s/csv/top_sus_%s.csv' % (
        project, project)
    # 分了是否为真实i缺陷-start
    isFaultList = []
    noFaultList = []
    csvLine = 1  # csv行
    with open(CSV_PATH) as f:
        csvFile = csv.reader(f)
        for row in csvFile:
            if csvLine != 1:
                if str(row[19]) == str(1):
                    isFaultList.append(row)
                else:
                    noFaultList.append(row)
            csvLine += 1
    # 分了是否为真实i缺陷-end
    if miniK == -1:
        if len(isFaultList) >= 10:
            k = 10  # 默认10次交叉验证或者更少
        else:
            # k = len(isFaultList)
            k = len(isFaultList)
    else:
        k = miniK
    isFault_step = math.floor(len(isFaultList)/k)  # 交叉验证，真实缺陷个数（步长）
    noFault_step = math.floor(len(noFaultList)/k)  # 交叉验证，非真实缺陷个数（步长）
    # 文件保存父路径
    SAVE_PATH = '/Users/lvlaxjh/code/CBFL/data/%s/csv/traintest/' % (project)
    isF_flag_TEM = 0
    noF_flag_TEM = 0
    for i in range(k):
        trainList_dataset = []
        testList_dataset = []
        # 根据步长和切割生成测试集-start
        for j in isFaultList[isF_flag_TEM:isF_flag_TEM+isFault_step]:
            # TEM_j = j
            # TEM_j.append(-1)
            testList_dataset.append(j)  # 真实缺陷
        isF_flag_TEM = isF_flag_TEM+isFault_step
        for j in noFaultList[noF_flag_TEM:noF_flag_TEM+noFault_step]:
            # TEM_j = j
            # TEM_j.append(-1)
            testList_dataset.append(j)  # 非真实缺陷
        noF_flag_TEM = noF_flag_TEM+noFault_step
        # 根据步长和切割生成测试集-end
        # 根据剩下数据生成训练集-start
        for j in isFaultList:
            if j not in testList_dataset:
                trainList_dataset.append(j)
        for j in noFaultList:
            if j not in testList_dataset:
                trainList_dataset.append(j)
        # 根据剩下数据生成训练集-end
        csvFiletest = open(SAVE_PATH+'%s-test-%s.csv' %
                           (project, str(i)), 'w', encoding="utf-8", newline="")
        csvWritertest = csv.writer(csvFiletest)
        csvWritertest.writerow(['dataset', 'project', 'path', 'version', 'codeLine', 'statement', 'varTotal', 'optTotal', 'array', 'bracketDepth',
                                'bracketTotal', 'keywordTotal', 'methodTotal', 'typeTotal', 'logic', 'lengthEle', 'lengthWord', 'depth', 'suspicious', 'accuracy', 'predict'])
        for n in testList_dataset:
            TEM_n = copy.deepcopy(n)
            TEM_n.append(-1)
            csvWritertest.writerow(TEM_n)
        csvFiletest.close()
        csvFiletrain = open(SAVE_PATH+'%s-train-%s.csv' %
                            (project, str(i)), 'w', encoding="utf-8", newline="")
        csvWritertrain = csv.writer(csvFiletrain)
        csvWritertrain.writerow(['dataset', 'project', 'path', 'version', 'codeLine', 'statement', 'varTotal', 'optTotal', 'array', 'bracketDepth',
                                 'bracketTotal', 'keywordTotal', 'methodTotal', 'typeTotal', 'logic', 'lengthEle', 'lengthWord', 'depth', 'suspicious', 'accuracy'])

        for m in trainList_dataset:
            csvWritertrain.writerow(m)
        csvFiletrain.close()
# 切割数据集用于交叉验证-end


# # 将全部真实缺陷（无论排名）存储（非必须使用，只针对真实缺陷过少情况）-start
# def get_all_true_sus(version,project):
#     top_sus_timeList = []
#     with open('/Users/lvlaxjh/code/CBFL/data/%s/csv/top_sus_%s.csv' % (
#         project, project)) as f:
#         csvFile = csv.reader(f)
#         csvLine = 1  # csv行
#         # 取列表中最大可疑度-start
#         for i in sortRow:
#             if topSus == float(i[18]):
#                 topSusList.append(i)
#         # 取列表中最大可疑度-end
#     print('get_sus_top %s%s.csv -> ok' % (project, str(version)))


if __name__ == "__main__":
    project = 'mockito'
    k = 10
    topSusList = []
    for i in range(1, 39):  # (1，项目总数+1)
        # if i != 2:
            # CODE_FILE_PATH = '/Volumes/Elements/DSatr/%s-direct/%s-%s/source/' % (
            #     project, project, str(i))  # java文件父目录
        CODE_FILE_PATH = '/Volumes/Elements/DSatr/%s-direct/%s-%s/src/' % (
            project, project, str(i))  # java文件父目录
        # ----------------------------------------------------------------
        # get_project_sus(i, project, CODE_FILE_PATH)# 找到真实缺陷语句并存储excel,in for
        # print(f'\033[0;36;40m\t step1 get_project_sus {project} {str(i)} -> ok \033[0m')
        # ----------------------------------------------------------------
        run_java_get_ASTjson(i, project, CODE_FILE_PATH)# 生成对应文件AST树,in for
        print(
            f'\033[0;36;40m\t step2 run_java_get_ASTjson {project} {str(i)} -> ok \033[0m')
        analytic_complexity.save_as_csv('/Users/lvlaxjh/code/CBFL/data/%s/%s-%s.xlsx' % (project, project, str(i)),
                                        '/Users/lvlaxjh/code/CBFL/data/%s/csv/' % (
                                            project),
                                        CODE_FILE_PATH,
                                        i, project)  # 计算特征值，并存储,in for
        print(f'\033[0;36;40m\t step3 ac.save_as_csv {project} {str(i)} -> ok \033[0m')
        #----------------------------------------------------------------
            # set_sus_one(i, project)  # 将可疑度大与1的设置为1
            # print(f'\033[0;36;40m\t step4 set_sus_one {project} {str(i)} -> ok \033[0m')
            # ----------------------------------------------------------------
            # 获取可疑度并列排名第一的语句,in for,-dont cut
    #         get_sus_top(i, project, topSusList, True)
    #         print(
    #             f'\033[0;36;40m\t step5 get_sus_top {project} {str(i)} -> ok \033[0m')
    # save_top_sus(topSusList, project)  # 保存全部定并列第一文件,-dont cut
    # print(f'\033[0;36;40m\t step6 get_sus_top {project} -> ok \033[0m')
    # ----------------------------------------------------------------
    # cut_data_for_test_train(project, k)  # 切割数据集用于交叉验证,当数据量小于10时，设置的k折
    # print(f'\033[0;36;40m\t step7 cut_data_for_test_train {project} -> ok \033[0m')
