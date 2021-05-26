'''
Author: your name
Date: 2021-05-26 14:32:36
LastEditTime: 2021-05-26 16:15:49
Description: 计算可疑度
'''
import openpyxl
pid = 27
MAIN_PATH = '/Users/lvlaxjh/code/dataset/Time/gzoltars/Time/%s/' % (
    str(pid))  # 包含matrix，spectra的父路径
CREATE_EXCEL_PATH = '/Users/lvlaxjh/code/CBFL/data/excel/'  # 创建excel父路径
BUGGY_LINE_PATH = '/Users/lvlaxjh/code/CBFL/data/excel/Time-%s.buggy.lines' % (
    str(pid))  # 真实缺陷位置文件buggy.lines的绝对路径
# /Users/lvlaxjh/code/dataset/Time/gzoltars/Time/27/matrix
# matrix，spectra
'''
ef = - & 1个数
ep = + & 1个数
nf = - & 0个数
(ef*ef)/(ep+nf)
'''
ef = 0
ep = 0
nf = 0
file = open(MAIN_PATH+'matrix', 'r')
fileLineList = file.readlines()
# print(fileLineList[0])
column = len(fileLineList[0].rstrip("\n").split(' '))-1
codeLinesusList = []
for j in range(column):
    for i in fileLineList:
        flag = i.rstrip("\n").split(' ')[-1]
        num = i.rstrip("\n").split(' ')[j]
        if num == '1' and flag == '-':
            ef += 1
        if num == '1' and flag == '+':
            ep += 1
        if num == '0' and flag == '-':
            nf += 1
    sus = 0
    if (ep+nf) == 0:
        sus = 0
    else:
        sus = (ef*ef)/(ep+nf)
    ef = 0
    ep = 0
    nf = 0
    codeLinesusList.append(sus)
    print(sus)
print(codeLinesusList)
print(len(codeLinesusList))
file.close()

file = open(MAIN_PATH+'spectra', 'r')
buggyFile = open(BUGGY_LINE_PATH,'r')
buggyFileLineList = buggyFile.readlines()
# buggyLineList = []
# for i in buggyFileLineList:
#     buggyLineList.append(i.split('#')[1])
fileLineList = file.readlines()
excel = openpyxl.Workbook()
sheet = excel.worksheets[0]  # 表
sheetHead = ['项目路径', '行数', 'statement', 'dstar', 'faulty']
for i in range(len(sheetHead)):
    sheet.cell(1, i+1).value = sheetHead[i]
for i in range(5947):
    for j in buggyFileLineList:
        if j.split('#')[0][:-5] == fileLineList[i].split('#')[0] and j.split('#')[1] == fileLineList[i].split('#')[1]:
            sheet.cell(i+1, 1).value = fileLineList[i].split('#')[0]
            sheet.cell(i+1, 2).value = fileLineList[i].split('#')[1]
            sheet.cell(i+1, 4).value = codeLinesusList[i]
            sheet.cell(i+1, 5).value = 1
        else:
            sheet.cell(i+1, 1).value = fileLineList[i].split('#')[0]
            sheet.cell(i+1, 2).value = fileLineList[i].split('#')[1]
            sheet.cell(i+1, 4).value = codeLinesusList[i]
            sheet.cell(i+1, 5).value = 0
excel.save(CREATE_EXCEL_PATH+str(pid)+'.xlsx')
file.close()
buggyFile.close()
