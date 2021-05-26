import openpyxl
import shutil
import os
CODE_FILE_PATH= r'D:\GithubDeskTop\CBFL\Data\lang-direct\lang-direct\lang-37\src\java'
EXCEL_FILE_PATH = r'D:\GithubDeskTop\CBFL\Data\excels_lang\MUSE\lang37.xlsx'
Target_FILE_PATH = r'D:\GithubDeskTop\CBFL\Data\excels_lang\MUSE\lang-37'

# create the folders if not already exists
os.makedirs(Target_FILE_PATH)
excel = openpyxl.load_workbook(EXCEL_FILE_PATH)#excel文件
sheet = excel.worksheets[0]#表
line = 2 #sheet 行
codeLineList = []#存储代码行
#读sheet-start
# while True:
#     if sheet.cell(line, 2).value != None:
#         codeLineList.append(sheet.cell(line, 2).value)
#     else:
#         break
#     line+=1
# line = 2
#读sheet-end
#读代码，写sheet-start
while True:
    if sheet.cell(line, 2).value != None:
        the_line = CODE_FILE_PATH+"\\"+ str(sheet.cell(line, 1).value).replace(".", "\\")+'.java'
        try:
            shutil.copy(the_line, Target_FILE_PATH)
        except IOError as e:
            print("Unable to copy file. %s" % e)
        except:
            print("Unexpected error:")
    else:
        break
    line +=1


