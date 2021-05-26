'''
Author: your name
Date: 2021-05-18 09:51:21
LastEditTime: 2021-05-26 15:48:29
Description: file content
'''
import linecache
import openpyxl
CODE_FILE_PATH = r'F:\崔老师组会文件\缺陷定位工作\源代码缺陷定位\time-direct\time-direct\time-17\src\main\java'
EXCEL_FILE_PATH = r'F:\崔老师组会文件\缺陷定位工作\源代码缺陷定位\excels_time\time17.xlsx'
excel = openpyxl.load_workbook(EXCEL_FILE_PATH)#excel文件
sheet = excel.worksheets[0]#表
line = 2#sheet 行
codeLineList = []#存储代码行
#读sheet-start
while True:
    if sheet.cell(line, 2).value != None:
        codeLineList.append(sheet.cell(line, 2).value)
    else:
        break
    line+=1
line = 2
#读sheet-end
#读代码，写sheet-start
for i in codeLineList:
    the_line = linecache.getline(
        CODE_FILE_PATH+"\\"+ str(sheet.cell(line, 1).value).replace(".", "\\")+'.java', i)
    print(the_line.rstrip("\n"))
    sheet.cell(line, 3).value = the_line.rstrip("\n")
    line +=1
#读代码，写sheet-end
excel.save(EXCEL_FILE_PATH)#存储文件

