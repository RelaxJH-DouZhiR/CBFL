'''
Author: your name
Date: 2021-05-28 08:43:20
LastEditTime: 2021-05-28 09:58:16
Description: 将不同版本的数据集合为excel
'''

import openpyxl
FATHER_PATH = '/Users/lvlaxjh/code/dataset/Data/excels_chart/DStar/'
TARGET_EXCEL_PATH = '/Users/lvlaxjh/code/CBFL/data/'
EXCEL_NAME = 'chart.xlsx'
FILE_LIST_VID = [3, 8, 9, 10, 11, 13, 16, 18, 19, 24]
contentList = []

for i in FILE_LIST_VID:
    # vId = i
    excel = openpyxl.load_workbook(
        FATHER_PATH+'chart-%s/chart%s.xlsx' % (str(i), str(i)))  # 打开excel
    sheet = excel.worksheets[0]  # 表
    sheetLine = 2  # sheet 行
    while True:
        if sheet.cell(sheetLine, 1).value != None:
            contentList.append({'path': sheet.cell(sheetLine, 1).value,
                                'line': sheet.cell(sheetLine, 2).value,
                                'statement': sheet.cell(sheetLine, 3).value,
                                'dstar': sheet.cell(sheetLine, 4).value,
                                'faulty': sheet.cell(sheetLine, 5).value,
                                'vid': i})
            sheetLine += 1
        else:
            break
    excel.close()

print(len(contentList))
tarExcel = openpyxl.Workbook()
tarSheet = tarExcel.worksheets[0]  # 表
sheetHead = ['项目路径', '行数', 'statement', 'dstar', 'faulty', 'pid']
for i in range(len(sheetHead)):
    tarSheet.cell(1, i+1).value = sheetHead[i]
sheetLine = 2  # sheet 行
for i in contentList:
    tarSheet.cell(sheetLine, 1).value = i['path']
    tarSheet.cell(sheetLine, 2).value = i['line']
    tarSheet.cell(sheetLine, 3).value = i['statement']
    tarSheet.cell(sheetLine, 4).value = i['dstar']
    tarSheet.cell(sheetLine, 5).value = i['faulty']
    tarSheet.cell(sheetLine, 6).value = i['vid']
    sheetLine += 1
# sheetLine = 2
tarExcel.save(TARGET_EXCEL_PATH+EXCEL_NAME)
# print(contentList)
