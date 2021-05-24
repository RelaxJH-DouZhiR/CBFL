'''
Author: your name
Date: 2021-05-24 08:42:18
LastEditTime: 2021-05-24 13:35:51
Description: file content
'''
import sys
import openpyxl
# import shutil
import os
# CODE_FILE_PATH = r'D:\GithubDeskTop\CBFL\Data\lang-direct\lang-direct\lang-63\src\java'
APP_JAVA = '/Users/lvlaxjh/code/CBFL/AST/EclipseAST/App.java'
MAIN_PATH = '/Users/lvlaxjh/code/dataset/d4j/lang_1_buggy/src/main/java'
EXCEL_FILE_PATH = '/Users/lvlaxjh/code/CBFL/Data/excels_lang/DStar/lang-1/lang1.xlsx'
JAVA_FILE_PATH = ''
# /Users/lvlaxjh/code/dataset/d4j/lang_1_buggy/src/main/java/org/apache/commons/lang3/
JAVA_FILE = ''
# StringUtils

excel = openpyxl.load_workbook(EXCEL_FILE_PATH)  # excel文件
sheet = excel.worksheets[0]  # 表
line = 2  # sheet 行
javaFileList = []
javaFileNames = []
while True:
    if sheet.cell(line, 1).value != None:
        filePath = str(sheet.cell(line, 1).value).replace(".", "/")
        if filePath not in javaFileList:
            javaFileNames.append(filePath.split('/')[-1])
            javaFileList.append(filePath)
    else:
        break
    line += 1

for i in range(len(javaFileList)):
    # print(javaFileList)
    paths = ''
    for j in javaFileList[i].split('/')[:-1]:
        paths += j
        paths += '/'
    JAVA_FILE_PATH = MAIN_PATH+'/'+paths
    JAVA_FILE = javaFileNames[i]
    file = open(APP_JAVA, 'r+')
    fileList = file.readlines()
    fileList[51] = '		'+'String MAIN_PATH = \"'+JAVA_FILE_PATH+'\";\n'
    fileList[52] = '		'+'String JAVA_FILE = \"'+JAVA_FILE+'\";\n'
    file.close()
    file = open(APP_JAVA, 'w+')
    file.writelines(fileList)
    file.close()
    os.popen("cd /Users/lvlaxjh/code/CBFL/AST/EclipseAST")
    osres = os.popen(
        'javac -Xlint:deprecation -cp \'lib/*\' *.java && java -Xss4m -cp .:\'lib/*\' App "$@"')
    print(osres.read())
