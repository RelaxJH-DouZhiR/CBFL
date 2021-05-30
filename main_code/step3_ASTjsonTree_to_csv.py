'''
Author: your name
Date: 2021-05-20 14:47:36
LastEditTime: 2021-05-28 14:51:14
Description: 根据excel生成复杂度数据集csv
'''
import json
import copy
import re
import csv
import openpyxl
import shutil
import sys
import openpyxl
import os
ids = "17"
projects = "time"  # ***
GLOBAL_VAR = {
    # AST中关键字
    # AST关键字
    "AST_KEYWORD_LIST": ["identifier", "value", "keyword", "escapedValue", "operator", "booleanValue", "expression"],
    # 原java文件的路径 ***
    # "CODE_MAIN_PATH": "/Users/lvlaxjh/code/dataset/Data/excels_%s/DStar/%s-%s" % (projects, projects, ids),
    "CODE_MAIN_PATH": "/Users/lvlaxjh/code/dataset/Data/excels_%s/DStar/" % (projects),
    # 读取的excel文件路径 ***
    # "EXCEL_PATH": "/Users/lvlaxjh/code/CBFL/Data/excels_%s/DStar/%s-%s/%s%s.xlsx" % (projects, projects, ids, projects, ids),
    "EXCEL_PATH": "/Users/lvlaxjh/code/CBFL/data/time/time.xlsx",
    "P_NUMBER": ids,  # 编号 ***
    "project": projects,  # 项目 ***
    "JAVA_PATH": "",
    "JSON_PATH": "",  # AST树json文件路径
    # 数据集csv路径
    "CSV_PATH": "/Users/lvlaxjh/code/CBFL/data/%s/%s_data.csv" % (projects, projects),
    "ASTDict": {},  # AST树-json转为字典
    "resultDict": [],  # 结果字典
    #
    'AST_TO_JSON_APP_JAVA': '/Users/lvlaxjh/code/CBFL/AST/EclipseAST/App.java',
    'EclipseAST': '/Users/lvlaxjh/code/CBFL/AST/EclipseAST',
    # 'JAVA_FILE_PATH' : '',
    # 'JAVA_FILE' : '',


}


# AST包，库，类遍历-start
def get_code_element(targetCodeLine, codeStatement):  # 目标代码行数，代码内容
    global GLOBAL_VAR
    file = open(GLOBAL_VAR['JSON_PATH'], 'r')
    ASTDict = json.loads(file.read())
    for key, val in ASTDict.items():
        # 包-start
        if key == "package":  # 包-AST
            packageItem = val
            get_code_recursion(packageItem, [], targetCodeLine, codeStatement)
        # 包-end
        # 库-start
        if key == "imports":
            importItem = val
            for i in importItem:
                get_code_recursion(i, [], targetCodeLine, codeStatement)
        # 库-end
        # 类-start
        if key == "types":  # 类-AST
            typeItem = val
            for i in typeItem:
                get_code_recursion(i, [], targetCodeLine, codeStatement)
        # 类-end
    # print(GLOBAL_VAR['resultDict'])
# AST包，库，类遍历-end


# 递归遍历AST树-start
def get_code_recursion(recItem, nodeStack, targetCodeLine, codeStatement):  # 树，节点栈，目标代码行数，代码内容
    global GLOBAL_VAR
    if type(recItem) is dict:  # 字典递归
        if 'node' in recItem.keys():
            nodeStack.append(recItem['node'])  # 父节点入栈
        if 'node' in recItem.keys() and recItem['node_line'] == targetCodeLine:
            for i in GLOBAL_VAR['AST_KEYWORD_LIST']:
                if i in recItem.keys() and str(recItem[i]) in codeStatement:
                    nodeDice_TEM = {
                        "code": recItem[i],
                        "code_line": recItem['node_line'],
                        "father_node": copy.deepcopy(nodeStack),
                    }
                    GLOBAL_VAR['resultDict'].append(nodeDice_TEM)
        for key, val in recItem.items():
            if type(val) is dict or type(val) is list:  # 深层递归
                get_code_recursion(
                    val, nodeStack, targetCodeLine, codeStatement)
        if 'node' in recItem.keys():  # 父节点出栈
            nodeStack.pop()
    if type(recItem) is list:  # 列表递归
        for i in recItem:
            get_code_recursion(i, nodeStack, targetCodeLine, codeStatement)
# 递归遍历AST树-end


# 获取括号数量、嵌套层数-start
def get_brackets_nesting(codeStatement):  # 代码内容（返回：括号嵌套层数，括号总数）
    global GLOBAL_VAR
    bracketsStack = []  # 括号栈
    depth = 0  # 括号深度
    total = 0  # 括号数量
    for i in codeStatement:
        if i == '(':
            bracketsStack.append(')')
        elif i == '[':
            bracketsStack.append(']')
        elif i == '{':
            bracketsStack.append('}')
        if len(bracketsStack) != 0 and (i == ')' or i == ']' or i == '}'):  # 栈不为空
            bracketsStack.pop()
            depth += 1
        # 计算总数
        if i == '(' or i == ')' or i == '[' or i == ']' or i == '{' or i == '}':
            total += 1
    return depth, total
# 获取括号数量、嵌套层数-end


# 逻辑-start
def get_logic_word(codeStatement):  # 代码内容（返回：bool是否包含逻辑）
    global GLOBAL_VAR
    logicList = ['if', 'else if', 'else', 'switch', 'case', 'while', 'for']
    for n in logicList:
        pattern = re.compile(n)   # 模式
        for i in pattern.finditer(codeStatement):
            if i.start() > 0 and (codeStatement[i.start()-1].isdigit() or codeStatement[i.start()-1].isalpha()):
                return False
            elif (i.start()+len(n)) > len(n) or ((i.start()+len(n)) < len(n) and codeStatement[i.start()+len(n)].isdigit() or codeStatement[i.start()+len(n)].isalpha()):
                return False
            else:
                return True
# 逻辑-end


# 关键字-start
def get_keyword(codeStatement):  # 代码内容（返回：关键字数量）
    global GLOBAL_VAR
    keywordList = ['return']
    total = 0
    for n in keywordList:
        pattern = re.compile(n)   # 模式
        for i in pattern.finditer(codeStatement):
            if i.start() > 0 and (codeStatement[i.start()-1].isdigit() or codeStatement[i.start()-1].isalpha()):
                pass
            elif (i.start()+len(n)) > len(n) or ((i.start()+len(n)) < len(n) and codeStatement[i.start()+len(n)].isdigit() or codeStatement[i.start()+len(n)].isalpha()):
                pass
            else:
                total += 1
    return total
# 关键字-end


# csv存储结构-start
def get_data_for_csv(resultDict, targetCodeLine, codeStatement):  # resultDict,目标代码行数，代码内容
    global GLOBAL_VAR
    csvRes = {
        'varTotal': 0,  # 变量+
        'optTotal': 0,  # 运算符+
        'array': 0,  # 数组-01+
        'bracketDepth': 0,  # 括号深度+
        'bracketTotal': 0,  # 括号总数+
        'keywordTotal': 0,  # 关键字+
        'methodTotal': 0,  # 函数+
        'typeTotal': 0,  # 类+
        'logic': 0,  # 逻辑，01+
        'lengthEle': 0,  # 元素数+
        'lengthWord': 0,  # 字数+
        'depth': 0,  # 深度+
    }
    csvRes['lengthWord'] = len(codeStatement.strip())  # 字数
    csvRes['lengthEle'] = len(resultDict)  # 元素数
    # 深度计算-start
    depth_TEM = 0
    depth_TEM = len(resultDict[0]['father_node'])
    for i in resultDict:
        if len(i['father_node']) < depth_TEM:
            depth_TEM = len(i['father_node'])
    flag_TEM = 0
    while True:
        compareFlag_TEM = ''
        canBreak_TEM = False
        for i in range(len(resultDict)):
            if i == 0:
                compareFlag_TEM = resultDict[i]['father_node'][flag_TEM]
            else:
                if resultDict[i]['father_node'][flag_TEM] != compareFlag_TEM:
                    canBreak_TEM = True
        if flag_TEM == depth_TEM-1:
            canBreak_TEM = True
        flag_TEM += 1
        if canBreak_TEM:
            csvRes['depth'] = flag_TEM-1  # 存储深度
            break
    # 深度计算-end
    if get_logic_word(codeStatement):  # 逻辑
        csvRes['logic'] = 1
    csvRes['bracketDepth'], csvRes['bracketTotal'] = get_brackets_nesting(
        codeStatement)
    csvRes['keywordTotal'] += get_keyword(codeStatement)
    for i in resultDict:
        if i['father_node'][len(i['father_node'])-1] == 'InfixExpression':  # 运算符
            csvRes['optTotal'] += 1
            continue
        if i['father_node'][len(i['father_node'])-1] == 'Modifier':  # 关键字
            csvRes['keywordTotal'] += 1
            continue
        if i['father_node'][len(i['father_node'])-1] == 'SimpleName' and i['father_node'][len(i['father_node'])-2] == 'SimpleType':  # 类
            csvRes['typeTotal'] += 1
            continue
        if i['father_node'][len(i['father_node'])-1] == 'SimpleName' and i['father_node'][len(i['father_node'])-2] == 'MethodInvocation':  # 函数
            csvRes['methodTotal'] += 1
            continue
        if i['father_node'][len(i['father_node'])-1] == 'SimpleName' and (i['father_node'][len(i['father_node'])-2] == 'ArrayCreation' or i['father_node'][len(i['father_node'])-2] == 'ArrayAccess'):  # 数组
            csvRes['array'] = 1
            continue
        if i['father_node'][len(i['father_node'])-1] == 'SimpleName' or i['father_node'][len(i['father_node'])-1] == 'booleanValue':  # 变量,优先级最低
            csvRes['varTotal'] += 1
    return csvRes
# csv存储结构-end


# 运行java生成json-start
def run_java():
    global GLOBAL_VAR
    excel = openpyxl.load_workbook(GLOBAL_VAR['EXCEL_PATH'])  # excel文件
    sheet = excel.worksheets[0]  # 表
    line = 2  # sheet 行
    javaFileList = []
    javaFileNames = []
    while True:
        if sheet.cell(line, 1).value != None:
            filePath = str(sheet.cell(line, 1).value).replace(".", "/")
            if filePath not in javaFileList:
                javaFileNames.append(filePath.split('/')[-1])
                javaFileList.append(filePath)
        else:
            break
        line += 1

    for i in range(len(javaFileList)):
        paths = ''
        for j in javaFileList[i].split('/')[:-1]:
            paths += j
            paths += '/'
        JAVA_FILE_PATH = GLOBAL_VAR['CODE_MAIN_PATH']+'/'+paths
        JAVA_FILE = javaFileNames[i]
        file = open(GLOBAL_VAR['AST_TO_JSON_APP_JAVA'], 'r+')
        fileList = file.readlines()
        fileList[51] = '		'+'String MAIN_PATH = \"'+JAVA_FILE_PATH+'\";\n'
        fileList[52] = '		'+'String JAVA_FILE = \"'+JAVA_FILE+'\";\n'
        file.close()
        file = open(GLOBAL_VAR['AST_TO_JSON_APP_JAVA'], 'w+')
        file.writelines(fileList)
        file.close()
        # os.popen("cd /Users/lvlaxjh/code/CBFL/AST/EclipseAST")
        osres = os.popen(
            'cd '+GLOBAL_VAR['EclipseAST']+' && javac -Xlint:deprecation -cp \'lib/*\' *.java && java -Xss4m -cp .:\'lib/*\' App "$@"')
        print(osres.read())
# 运行java生成json-end


if __name__ == "__main__":
    #
    # run_java()
    #
    excel = openpyxl.load_workbook(GLOBAL_VAR["EXCEL_PATH"])  # 加载excel
    sheet = excel.worksheets[0]  # 表
    excelLine = 2  # excel行
    #
    codeLine = 0  # 代码行数
    codeStatement = ""  # 代码
    suspicious = 0  # 可疑度
    accuracy = 0  # 是否为真实缺陷
    pid = 0  # 项目id
    #
    shutil.copyfile(GLOBAL_VAR['CSV_PATH'],
                    GLOBAL_VAR['CSV_PATH'][:-4]+'_backups.csv')  # 备份csv
    #
    print(" --- csv start ---")
    csvFile = open(GLOBAL_VAR['CSV_PATH'], 'w', encoding="utf-8", newline="")
    csvWriter = csv.writer(csvFile)
    csvWriter.writerow(['dataset', 'project', 'pPath', 'pId', 'codeLine', 'statement', 'varTotal', 'optTotal', 'array', 'bracketDepth',
                                   'bracketTotal', 'keywordTotal', 'methodTotal', 'typeTotal', 'logic', 'lengthEle', 'lengthWord', 'depth', 'suspicious', 'accuracy'])

    while True:
        if sheet.cell(excelLine, 1).value != None:
            codeLine = sheet.cell(excelLine, 2).value
            codeStatement = sheet.cell(excelLine, 3).value
            suspicious = sheet.cell(excelLine, 4).value
            accuracy = sheet.cell(excelLine, 5).value
            pid = sheet.cell(excelLine, 6).value
            GLOBAL_VAR["JAVA_PATH"] = GLOBAL_VAR["CODE_MAIN_PATH"] + '%s-%s/' % (
                projects, pid)+str(sheet.cell(excelLine, 1).value).replace(".", "/")+'.java'
            GLOBAL_VAR["JSON_PATH"] = GLOBAL_VAR["CODE_MAIN_PATH"] + '%s-%s/' % (
                projects, pid)+str(sheet.cell(excelLine, 1).value).replace(".", "/")+'.json'
            #
            get_code_element(codeLine, codeStatement)  # 解析
            if GLOBAL_VAR['resultDict'] == []:
                print('''\033[31m
<------error------>
java_path -> %s
json_path -> %s
codeLine -> %s
codeStatement -> %s
<------error------>
\033[0m''' % (GLOBAL_VAR["JAVA_PATH"], GLOBAL_VAR["JSON_PATH"], str(codeLine), codeStatement))
            else:
                astResultDict = get_data_for_csv(
                    GLOBAL_VAR['resultDict'], codeLine, codeStatement)

                # csvWriter = csv.writer(csvFile)

                csvWriter.writerow(['defect4j', GLOBAL_VAR['project'], str(sheet.cell(excelLine, 1).value).replace(".", "/"), pid, codeLine, codeStatement, astResultDict['varTotal'], astResultDict['optTotal'], astResultDict['array'], astResultDict['bracketDepth'],
                                    astResultDict['bracketTotal'], astResultDict['keywordTotal'], astResultDict['methodTotal'], astResultDict['typeTotal'], astResultDict['logic'], astResultDict['lengthEle'], astResultDict['lengthWord'], astResultDict['depth'], suspicious, accuracy])
                GLOBAL_VAR['resultDict'] = []
            #
            excelLine += 1

        else:
            print(" --- csv end ---")
            break
    #
    csvFile.close()
