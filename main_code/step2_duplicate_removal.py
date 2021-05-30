'''
Author: your name
Date: 2021-05-28 08:43:20
LastEditTime: 2021-05-30 21:27:24
Description: 数据去重
'''

import openpyxl
import csv

EXCEL_PATH = '/Users/lvlaxjh/code/CBFL/data/chart/chart_datatest.xlsx'
TARGET_CSV_PATH = '/Users/lvlaxjh/code/CBFL/data/chart/chart_datatest1.csv'

contentList = []

excel = openpyxl.load_workbook(EXCEL_PATH)  # 打开excel
sheet = excel.worksheets[0]  # 表
sheetLine = 2  # sheet 行
while True:
    if sheet.cell(sheetLine, 1).value != None:
        comList = []
        for i in range(7,19):
            comList.append(sheet.cell(sheetLine, i).value)
        #
        isIns = True
        for i in contentList:
            if i['comList'] == comList:
                isIns = False
                print(comList)
        #
        if isIns:
            contentList.append({'dataset': sheet.cell(sheetLine, 1).value,
                                'project': sheet.cell(sheetLine, 2).value,
                                'pPath': sheet.cell(sheetLine, 3).value,
                                'pId': sheet.cell(sheetLine, 4).value,
                                'codeLine': sheet.cell(sheetLine, 5).value,
                                'statement': sheet.cell(sheetLine, 6).value,
                                'comList': comList,
                                'varTotal':sheet.cell(sheetLine, 7).value,
                                'optTotal':sheet.cell(sheetLine, 8).value,
                                'array':sheet.cell(sheetLine, 9).value,
                                'bracketDepth':sheet.cell(sheetLine, 10).value,
                                'bracketTotal':sheet.cell(sheetLine, 11).value,
                                'keywordTotal':sheet.cell(sheetLine, 12).value,
                                'methodTotal':sheet.cell(sheetLine, 13).value,
                                'typeTotal':sheet.cell(sheetLine, 14).value,
                                'logic':sheet.cell(sheetLine, 15).value,
                                'lengthEle':sheet.cell(sheetLine, 16).value,
                                'lengthWord':sheet.cell(sheetLine, 17).value,
                                'depth':sheet.cell(sheetLine, 18).value,
                                'suspicious':sheet.cell(sheetLine, 19).value,
                                'accuracy':sheet.cell(sheetLine, 20).value})
        isIns = True
        sheetLine += 1
    else:
        break
excel.close()

csvFile = open(TARGET_CSV_PATH, 'w', encoding="utf-8", newline="")
csvWriter = csv.writer(csvFile)
csvWriter.writerow(['dataset', 'project', 'pPath', 'pId', 'codeLine', 'statement', 'varTotal', 'optTotal', 'array', 'bracketDepth',
                    'bracketTotal', 'keywordTotal', 'methodTotal', 'typeTotal', 'logic', 'lengthEle', 'lengthWord', 'depth', 'suspicious', 'accuracy'])
for i in contentList:
    TEM_list = []
    for key,val in i.items():
        if key != 'comList':
            TEM_list.append(val)
    csvWriter.writerow(TEM_list)
csvFile.close()

# tarExcel = openpyxl.Workbook()
# tarSheet = tarExcel.worksheets[0]  # 表
# sheetHead = ['项目路径', '行数', 'statement', 'dstar', 'faulty', 'pid']
# for i in range(len(sheetHead)):
#     tarSheet.cell(1, i+1).value = sheetHead[i]
# sheetLine = 2  # sheet 行
# for i in contentList:
#     tarSheet.cell(sheetLine, 1).value = i['path']
#     tarSheet.cell(sheetLine, 2).value = i['line']
#     tarSheet.cell(sheetLine, 3).value = i['statement']
#     tarSheet.cell(sheetLine, 4).value = i['dstar']
#     tarSheet.cell(sheetLine, 5).value = i['faulty']
#     tarSheet.cell(sheetLine, 6).value = i['vid']
#     sheetLine += 1
# tarExcel.save(TARGET_EXCEL_PATH+EXCEL_NAME)
# # print(contentList)
# # for i in range(7,19):
# #     print(i)
