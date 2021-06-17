'''
Author: your name
Date: 2021-06-15 16:04:17
LastEditTime: 2021-06-17 16:02:48
Description: file content
'''
import json
from math import e
import openpyxl
import csv
from copy import deepcopy
FATHER_PATH = '/Users/lvlaxjh/code/'


# 获取原文件EXAM-start
def get_old_EXAM(project, version):
    EXCEL_PATH = f'{FATHER_PATH}CBFL/data/{project}/{project}-{version}.xlsx'
    excel = openpyxl.load_workbook(EXCEL_PATH)  # excel文件
    sheet = excel.worksheets[0]  # 表
    line = 2  # sheet 行
    EXAM_line = -1  # EXAM的行
    while True:
        if sheet.cell(line, 1).value != None:
            if sheet.cell(line, 6).value == 1 and EXAM_line == -1:
                EXAM_line = line
            #
            line += 1
        else:
            break
    if EXAM_line == -1:
        return 0
    else:
        return (EXAM_line-1)/(line-1)
# 获取原文件EXAM-end


# 获取新文件EXAM-start
def get_new_EXAM(project, version):  # input:项目,版本;return:EXAM值
    EXCEL_PATH = f'{FATHER_PATH}CBFL/data/{project}/{project}-{version}.xlsx'
    excel = openpyxl.load_workbook(EXCEL_PATH)  # excel文件
    sheet = excel.worksheets[0]  # 表
    line = 2  # sheet 行
    new_line = 2  # 新的行计算，去除所有预测为0的情况的
    EXAM_line = -1  # EXAM的行
    while True:
        if sheet.cell(line, 1).value != None:
            if sheet.cell(line, 6).value == 1 and EXAM_line == -1:
                EXAM_line = new_line
            if sheet.cell(line, 7).value == 1 or sheet.cell(line, 7).value == None:
                new_line += 1
            #
            line += 1
        else:
            break
    if EXAM_line == -1:
        return 0
    else:
        return (EXAM_line-1)/(line-1)
# 获取新文件EXAM-end


# 获取原文件Einspect-start
def get_old_Einspect(project, version, step):
    EXCEL_PATH = f'{FATHER_PATH}CBFL/data/{project}/{project}-{version}.xlsx'
    excel = openpyxl.load_workbook(EXCEL_PATH)  # excel文件
    sheet = excel.worksheets[0]  # 表
    line = 2  # sheet 行
    falut_sum = 0  # 错误总数
    while True:
        if sheet.cell(line, 1).value != None:
            if sheet.cell(line, 6).value == 1 and (line-1) <= step:
                falut_sum += 1
            #
            line += 1
        if (line+10) > step:
            break
        else:
            break
    return falut_sum
# 获取原文件Einspect-end


# 获取新文件Einspect-start
def get_new_Einspect(project, version, step):
    EXCEL_PATH = f'{FATHER_PATH}CBFL/data/{project}/{project}-{version}.xlsx'
    excel = openpyxl.load_workbook(EXCEL_PATH)  # excel文件
    sheet = excel.worksheets[0]  # 表
    line = 2  # sheet 行
    new_line = 2  # 新的行计算，去除所有预测为0的情况的
    falut_sum = 0  # 错误总数
    while True:
        if sheet.cell(line, 1).value != None:
            if sheet.cell(line, 6).value == 1 and (new_line-1) <= step:
                falut_sum += 1
            if sheet.cell(line, 7).value == 1 or sheet.cell(line, 7).value == None:
                new_line += 1
            #
            line += 1
        if (new_line+10) > step:
            break
        else:
            break
    return falut_sum
# 获取新文件Einspect-end

# 获取所有预测结果-start


def get_predict(project, percentageList, k):
    predictList = []
    predictDict = {
        'path': '',
        'version': '',
        'line': '',
        'percentage': '',
        'predict': '',
    }
    for id in range(k):
        for percentage in percentageList:
            with open(f'{FATHER_PATH}CBFL/data/{project}/csv/traintest/{percentage}-{project}-test-{str(id)}.csv') as f:
                csvFile = csv.reader(f)
                csvLine = 0  # csv行
                for row in csvFile:
                    if csvLine != 0:
                        predictDict['path'] = row[2]
                        predictDict['version'] = row[3]
                        predictDict['line'] = row[4]
                        predictDict['percentage'] = percentage
                        predictDict['predict'] = row[20]
                        predictList.append(deepcopy(predictDict))
                    csvLine += 1
    return predictList
# 获取所有预测结果-end


# 保存至源文件-start
def save_predict(project, versionList, percentageList, predictList, k):
    for version in versionList:
        EXCEL_PATH = f'{FATHER_PATH}CBFL/data/{project}/{project}-{version}.xlsx'
        excel = openpyxl.load_workbook(EXCEL_PATH)  # excel文件
        sheet = excel.worksheets[0]  # 表
        line = 2  # sheet 行
        sheet.cell(1, 7).value = 'pre_0.3'
        sheet.cell(1, 8).value = 'pre_0.5'
        sheet.cell(1, 9).value = 'pre_0.7'
        sheet.cell(1, 10).value = 'pre_1'
        sheet.cell(1, 11).value = 'pre_1.3'
        sheet.cell(1, 12).value = 'pre_1.5'
        sheet.cell(1, 13).value = 'pre_1.7'
        sheet.cell(1, 14).value = 'pre_2'
        while True:
            if sheet.cell(line, 1).value != None:
                for i in predictList:
                    if str(i['path']) == str(sheet.cell(line, 1).value).replace(".", "/") and str(i['version']) == str(sheet.cell(line, 2).value) and str(i['line']) == str(sheet.cell(line, 3).value):
                        print(f'{project}-{str(version)}-{str(line)} pre')
                        if i['percentage'] == 0.3:
                            sheet.cell(line, 7).value = i['predict']
                        if i['percentage'] == 0.5:
                            sheet.cell(line, 8).value = i['predict']
                        if i['percentage'] == 0.7:
                            sheet.cell(line, 9).value = i['predict']
                        if i['percentage'] == 1:
                            sheet.cell(line, 10).value = i['predict']
                        if i['percentage'] == 1.3:
                            sheet.cell(line, 11).value = i['predict']
                        if i['percentage'] == 1.5:
                            sheet.cell(line, 12).value = i['predict']
                        if i['percentage'] == 1.7:
                            sheet.cell(line, 13).value = i['predict']
                        if i['percentage'] == 2:
                            sheet.cell(line, 14).value = i['predict']
                #
                if line % 1000 == 0:
                    print(f'{project}-{str(version)}-{str(line)} pass ')
                line += 1
            else:
                break
        excel.save(EXCEL_PATH)
# 保存至源文件-end


if __name__ == "__main__":
    resultList = []
    resultDict = {
        'project': '',
        'old_EXAM':'',
        'new_EXAM':'',
        'old_Einspect': '',
        'new_Einspect':'',
    }
    # for project in ['chart', 'lang', 'time', 'math', 'closure', 'mockito']:
    for project in ['chart']:
        Einspect_step = 10  # *
        # project = 'mockito'  # *
        settingJson = open(f'{FATHER_PATH}CBFL/main_code/setting.json', 'r')
        settingContent = settingJson.read()
        setting = json.loads(settingContent)
        k = setting[project]['k']
        versionStart = setting[project]['versionStart']  # 开始版本
        versionEnd = setting[project]['versionEnd']  # 最后版本
        removeVersionList = setting[project]['removeVersionList']  # 不包含的版本
        percentageList = setting['percentageList']  # 前百分比数据列表
        topPercentageList = []  # 前百分比数据列表数据
        # 计算版本列表-start
        versionList = []  # 版本列表
        for i in range(versionStart, versionEnd+1):
            if i not in removeVersionList:
                versionList.append(i)
        # 计算版本列表-end
        average_EXAM_old = 0
        average_EXAM_new = 0
        sum_Einspect_old = 0
        sum_Einspect_new = 0
        EXAM_num = 0
        for i in versionList:
            old_EXAM = get_old_EXAM(project, i)
            print(
                f'\033[1;36m >{project}-{str(i)}< -old EXAM- {old_EXAM}  \033[0m')
            new_EXAM = get_new_EXAM(project, i)
            print(
                f'\033[1;36m >{project}-{str(i)}< -new EXAM- {new_EXAM}  \033[0m')
            if old_EXAM != new_EXAM:
                average_EXAM_old += old_EXAM
                average_EXAM_new += new_EXAM
                EXAM_num+=1
            # # ----------------------------------------------------------------
            old_Einspect = get_old_Einspect(project, i, 10)
            sum_Einspect_old += old_Einspect
            print(
                f'\033[1;35m >{project}-{str(i)}< -old Einspect@n-{Einspect_step}- {old_Einspect}  \033[0m')
            new_Einspect = get_new_Einspect(project, i, 10)
            sum_Einspect_new += new_Einspect
            print(
                f'\033[1;35m >{project}-{str(i)}< -new Einspect@n-{Einspect_step}- {new_Einspect}  \033[0m')
        # print(average_EXAM_old/len(versionList))
        # print(average_EXAM_new/len(versionList))
        resultDict['project'] = project
        resultDict['old_EXAM'] = average_EXAM_old/EXAM_num
        resultDict['new_EXAM'] = average_EXAM_new/EXAM_num
        resultList.append(deepcopy(resultDict))
        # print(sum_Einspect_old)
        # print(sum_Einspect_new)
        EXAM_num =0
        average_EXAM_old = 0
        average_EXAM_new = 0
        sum_Einspect_old = 0
        sum_Einspect_new = 0
        #
        EXCEL_PATH = f'{FATHER_PATH}CBFL/data/res.xlsx'
        excel = openpyxl.load_workbook(EXCEL_PATH)  # excel文件
        sheet = excel.worksheets[0]  # 表
        line = 2  # sheet 行
        print(resultList)
        for i in resultList:
            sheet.cell(line, 1).value = project
            sheet.cell(line, 2).value = i['old_EXAM']
            sheet.cell(line, 3).value = i['new_EXAM']
            line+=1
        excel.save(EXCEL_PATH)
