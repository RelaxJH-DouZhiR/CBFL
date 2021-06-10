'''
Author: your name
Date: 2021-06-02 16:15:11
LastEditTime: 2021-06-10 10:45:57
Description: 根据excel解析复杂度存储至csv
 █████╗ ███╗   ██╗ █████╗ ██╗  ██╗   ██╗████████╗██╗ ██████╗
██╔══██╗████╗  ██║██╔══██╗██║  ╚██╗ ██╔╝╚══██╔══╝██║██╔════╝
███████║██╔██╗ ██║███████║██║   ╚████╔╝    ██║   ██║██║
██╔══██║██║╚██╗██║██╔══██║██║    ╚██╔╝     ██║   ██║██║
██║  ██║██║ ╚████║██║  ██║███████╗██║      ██║   ██║╚██████╗
╚═╝  ╚═╝╚═╝  ╚═══╝╚═╝  ╚═╝╚══════╝╚═╝      ╚═╝   ╚═╝ ╚═════╝
 ██████╗ ██████╗ ███╗   ███╗██████╗ ██╗     ███████╗██╗  ██╗██╗████████╗██╗   ██╗
██╔════╝██╔═══██╗████╗ ████║██╔══██╗██║     ██╔════╝╚██╗██╔╝██║╚══██╔══╝╚██╗ ██╔╝
██║     ██║   ██║██╔████╔██║██████╔╝██║     █████╗   ╚███╔╝ ██║   ██║    ╚████╔╝
██║     ██║   ██║██║╚██╔╝██║██╔═══╝ ██║     ██╔══╝   ██╔██╗ ██║   ██║     ╚██╔╝
╚██████╗╚██████╔╝██║ ╚═╝ ██║██║     ███████╗███████╗██╔╝ ██╗██║   ██║      ██║
 ╚═════╝ ╚═════╝ ╚═╝     ╚═╝╚═╝     ╚══════╝╚══════╝╚═╝  ╚═╝╚═╝   ╚═╝      ╚═╝
'''
import json
import copy
import openpyxl
import shutil
import csv
import re


resultList = []  # 结果列表
JSON_PATH = '/Users/lvlaxjh/code/code/java/ExamManager/src/Exam.json'
# JSON_PATH = ''

# AST包，库，类遍历-start


def get_code_element(targetCodeLine, codeStatement):  # 目标代码行数，代码内容
    global JSON_PATH
    file = open(JSON_PATH, 'r')
    ASTDict = json.loads(file.read())
    for key, val in ASTDict.items():
        # 包-start
        if key == "package":  # 包-AST
            packageItem = val
            get_code_recursion(packageItem, [], targetCodeLine, codeStatement)
        # 包-end
        # 库-start
        if key == "imports":
            importItem = val
            for i in importItem:
                get_code_recursion(i, [], targetCodeLine, codeStatement)
        # 库-end
        # 类-start
        if key == "types":  # 类-AST
            typeItem = val
            for i in typeItem:
                get_code_recursion(i, [], targetCodeLine, codeStatement)
        # 类-end
# AST包，库，类遍历-end


# 递归遍历AST树-start
def get_code_recursion(recItem, nodeStack, targetCodeLine, codeStatement):  # 树，节点栈，目标代码行数，代码内容
    global resultList
    AST_KEYWORD_LIST = ["identifier", "value", "keyword",
                        "escapedValue", "operator", "booleanValue", "expression", 'identifier']
    if type(recItem) is dict:  # 字典递归
        if 'node' in recItem.keys():
            nodeStack.append(recItem['node'])  # 父节点入栈
        if 'node' in recItem.keys() and str(recItem['node_line']) == targetCodeLine:
            for i in AST_KEYWORD_LIST:
                if i in recItem.keys() and str(recItem[i]) in codeStatement:
                    nodeDice_TEM = {
                        "code": recItem[i],
                        "code_line": recItem['node_line'],
                        "father_node": copy.deepcopy(nodeStack),
                    }
                    resultList.append(nodeDice_TEM)
        for key, val in recItem.items():
            if type(val) is dict or type(val) is list:  # 深层递归
                get_code_recursion(
                    val, nodeStack, targetCodeLine, codeStatement)
        if 'node' in recItem.keys():  # 父节点出栈
            nodeStack.pop()
    if type(recItem) is list:  # 列表递归
        for i in recItem:
            get_code_recursion(i, nodeStack, targetCodeLine, codeStatement)
# 递归遍历AST树-end


# 获取括号数量、嵌套层数-start
def get_brackets_nesting(codeStatement):  # 代码内容（返回：括号嵌套层数，括号总数）
    bracketsStack = []  # 括号栈
    depth = 0  # 括号深度
    total = 0  # 括号数量
    if codeStatement != None:
        for i in codeStatement:
            if i == '(':
                bracketsStack.append(')')
            elif i == '[':
                bracketsStack.append(']')
            elif i == '{':
                bracketsStack.append('}')
            if len(bracketsStack) != 0 and (i == ')' or i == ']' or i == '}'):  # 栈不为空
                bracketsStack.pop()
                depth += 1
            # 计算总数
            if i == '(' or i == ')' or i == '[' or i == ']' or i == '{' or i == '}':
                total += 1
    return depth, total
# 获取括号数量、嵌套层数-end


# 逻辑-start
def get_logic_word(codeStatement):  # 代码内容（返回：bool是否包含逻辑）
    logicList = ['if', 'else if', 'else', 'switch', 'case', 'while', 'for']
    for n in logicList:
        pattern = re.compile(n)   # 模式
        if codeStatement != None:
            for i in pattern.finditer(codeStatement):
                if i.start() > 0 and (codeStatement[i.start()-1].isdigit() or codeStatement[i.start()-1].isalpha()):
                    return False
                elif (i.start()+len(n)) > len(n) or ((i.start()+len(n)) < len(n) and codeStatement[i.start()+len(n)].isdigit() or codeStatement[i.start()+len(n)].isalpha()):
                    return False
                else:
                    return True
        else:
            return False
# 逻辑-end


# 关键字-start
def get_keyword(codeStatement):  # 代码内容（返回：关键字数量）
    keywordList = ['return']
    total = 0
    for n in keywordList:
        pattern = re.compile(n)   # 模式
        if codeStatement != None:
            for i in pattern.finditer(str(codeStatement)):
                if i.start() > 0 and (codeStatement[i.start()-1].isdigit() or codeStatement[i.start()-1].isalpha()):
                    pass
                elif (i.start()+len(n)) > len(codeStatement) or ((i.start()+len(n)) < len(codeStatement) and codeStatement[i.start()+len(n)].isdigit() or codeStatement[i.start()+len(n)].isalpha()):
                    pass
                else:
                    total += 1
    return total
# 关键字-end


# csv存储结构-start
def get_data_for_csv(resultList, targetCodeLine, codeStatement):  # resultList,目标代码行数，代码内容
    csvRes = {
        'varTotal': 0,  # 变量+
        'optTotal': 0,  # 运算符+
        'array': 0,  # 数组-01+
        'bracketDepth': 0,  # 括号深度+
        'bracketTotal': 0,  # 括号总数+
        'keywordTotal': 0,  # 关键字+
        'methodTotal': 0,  # 函数+
        'typeTotal': 0,  # 类+
        'logic': 0,  # 逻辑，01+
        'lengthEle': 0,  # 元素数+
        'lengthWord': 0,  # 字数+
        'depth': 0,  # 深度+
    }
    csvRes['lengthWord'] = len(codeStatement.strip())  # 字数
    csvRes['lengthEle'] = len(resultList)  # 元素数
    # 深度计算-start
    depth_TEM = 0
    depth_TEM = len(resultList[0]['father_node'])
    for i in resultList:
        if len(i['father_node']) < depth_TEM:
            depth_TEM = len(i['father_node'])
    flag_TEM = 0
    while True:
        compareFlag_TEM = ''
        canBreak_TEM = False
        for i in range(len(resultList)):
            if i == 0:
                compareFlag_TEM = resultList[i]['father_node'][flag_TEM]
            else:
                if resultList[i]['father_node'][flag_TEM] != compareFlag_TEM:
                    canBreak_TEM = True
        if flag_TEM == depth_TEM-1:
            canBreak_TEM = True
        flag_TEM += 1
        if canBreak_TEM:
            csvRes['depth'] = flag_TEM-1  # 存储深度
            break
    # 深度计算-end
    if get_logic_word(codeStatement):  # 逻辑
        csvRes['logic'] = 1
    csvRes['bracketDepth'], csvRes['bracketTotal'] = get_brackets_nesting(
        codeStatement)
    csvRes['keywordTotal'] += get_keyword(codeStatement)
    for i in resultList:
        if i['father_node'][len(i['father_node'])-1] == 'InfixExpression':  # 运算符
            csvRes['optTotal'] += 1
            continue
        if i['father_node'][len(i['father_node'])-1] == 'Modifier':  # 关键字
            csvRes['keywordTotal'] += 1
            continue
        if i['father_node'][len(i['father_node'])-1] == 'SimpleName' and i['father_node'][len(i['father_node'])-2] == 'SimpleType':  # 类
            csvRes['typeTotal'] += 1
            continue
        if i['father_node'][len(i['father_node'])-1] == 'SimpleName' and i['father_node'][len(i['father_node'])-2] == 'MethodInvocation':  # 函数
            csvRes['methodTotal'] += 1
            continue
        if i['father_node'][len(i['father_node'])-1] == 'SimpleName' and (i['father_node'][len(i['father_node'])-2] == 'ArrayCreation' or i['father_node'][len(i['father_node'])-2] == 'ArrayAccess'):  # 数组
            csvRes['array'] = 1
            continue
        if i['father_node'][len(i['father_node'])-1] == 'SimpleName' or i['father_node'][len(i['father_node'])-1] == 'booleanValue':  # 变量,优先级最低
            csvRes['varTotal'] += 1
    return csvRes
# csv存储结构-end


# 计算复杂度并存储csv-start-------------------------------------------------------------------
def save_as_csv(EXCEL_PATH, CSV_FATHER_PATH, CODE_FATHER_PATH, version, project):
    global resultList, JSON_PATH
    '''
    EXCEL_PATH - excel路径
    CSV_FATHER_PATH - 存储文件csv的父路径
    JSON_PATH - ASTJSON的父路径，同代码父路径
    '''
    excel = openpyxl.load_workbook(EXCEL_PATH)  # 加载excel
    sheet = excel.worksheets[0]  # 表
    excelLine = 2  # excel行
    #
    codeLine = 0  # 代码行数
    codeStatement = ""  # 代码
    suspicious = 0  # 可疑度
    accuracy = 0  # 是否为真实缺陷
    #
    #
    csvFile = open(CSV_FATHER_PATH+'%s%s.csv' %
                   (project, str(version)), 'w', encoding="utf-8", newline="")
    csvWriter = csv.writer(csvFile)
    csvWriter.writerow(['dataset', 'project', 'path', 'version', 'codeLine', 'statement', 'varTotal', 'optTotal', 'array', 'bracketDepth',
                                   'bracketTotal', 'keywordTotal', 'methodTotal', 'typeTotal', 'logic', 'lengthEle', 'lengthWord', 'depth', 'suspicious', 'accuracy'])

    while True:
        if sheet.cell(excelLine, 1).value != None:
            ver = sheet.cell(excelLine, 2).value
            codeLine = sheet.cell(excelLine, 3).value
            codeStatement = sheet.cell(excelLine, 4).value
            suspicious = sheet.cell(excelLine, 5).value
            accuracy = sheet.cell(excelLine, 6).value
            # GLOBAL_VAR["JAVA_PATH"] = GLOBAL_VAR["CODE_MAIN_PATH"] + '%s-%s/' % (
            #     projects, pid)+str(sheet.cell(excelLine, 1).value).replace(".", "/")+'.java'
            JSON_PATH = CODE_FATHER_PATH + \
                str(sheet.cell(excelLine, 1).value).replace(".", "/")+'.json'
            # #
            if codeStatement != None:
                get_code_element(codeLine, codeStatement)  # 解析
                #AST无法解析 - start
                if resultList == []:
                    keywordTotal = get_keyword(codeStatement)  # 总数
                    if get_logic_word(codeStatement):
                        logic = 1
                    else:
                        logic = 0
                    bracketDepth, bracketTotal = get_brackets_nesting(
                        codeStatement)  # 括号深度，总数
                    csvWriter.writerow(['defect4j',
                                        project,
                                        str(sheet.cell(excelLine, 1).value).replace(
                                            ".", "/"),
                                        ver,
                                        codeLine,
                                        codeStatement,
                                        0,
                                        0,
                                        0,
                                        bracketDepth,
                                        bracketTotal,
                                        keywordTotal,
                                        0,
                                        0,
                                        logic,
                                        0,
                                        len(codeStatement.strip()),
                                        3,
                                        suspicious,
                                        accuracy])
                else:
                    astresultList = get_data_for_csv(
                        resultList, codeLine, codeStatement)
                    csvWriter.writerow(['defect4j', project, str(sheet.cell(excelLine, 1).value).replace(".", "/"),
                                        ver,
                                        codeLine,
                                        codeStatement,
                                        astresultList['varTotal'],
                                        astresultList['optTotal'],
                                        astresultList['array'],
                                        astresultList['bracketDepth'],
                                        astresultList['bracketTotal'],
                                        astresultList['keywordTotal'],
                                        astresultList['methodTotal'],
                                        astresultList['typeTotal'],
                                        astresultList['logic'],
                                        astresultList['lengthEle'],
                                        astresultList['lengthWord'],
                                        astresultList['depth'],
                                        suspicious,
                                        accuracy])
                    resultList = []
                #

                print('%s%s - %s -  %s -> ok' % (project, ver, str(sheet.cell(excelLine, 1).value).replace(
                    ".", "/"), str(codeLine)))
            excelLine += 1
        else:
            break
    #
    csvFile.close()
# 计算复杂度并存储csv-end-------------------------------------------------------------------
# if __name__ == "__main__":
#     get_code_element(14,'this.name=name;')
#     print(resultList)
