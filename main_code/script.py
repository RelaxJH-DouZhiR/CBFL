'''
Author: your name
Date: 2021-06-01 19:58:05
LastEditTime: 2021-06-02 21:02:44
Description: file content
'''
import openpyxl
import linecache
import os
import analytic_complexity


# 通过txt获取包含可疑度的excel-start
def get_project_sus(version, project):  # version为in,project为小写
    # 目标存储位置
    TARGET_EXCEL_PATH = '/Users/lvlaxjh/code/CBFL/data/%s/' % (project)
    # 包含真实缺陷
    TRUE_FAULT_FILE_PATH = '/Volumes/Elements/D4j/%s/%s/gzoltars/%s/%s/fault.txt' % (
        project.capitalize(), str(version), project.capitalize(), str(version))
    PROJECT_FILE_PATH = '/Volumes/Elements/D4j/%s/%s/gzoltars/%s/%s/Dstar_%s_%s.txt' % (
        project.capitalize(), str(version), project.capitalize(), str(version), project.capitalize(), str(version))  # 原文件
    #java路径需要修改
    CODE_FILE_PATH = '/Volumes/Elements/DSatr/%s-direct/%s-%s/source/' % (
        project, project, str(version))  # java文件父目录
    projectFile = open(PROJECT_FILE_PATH, 'r')  # 项目文件
    projectFileLineList = projectFile.readlines()
    trueFaultFile = open(TRUE_FAULT_FILE_PATH, 'r')  # 真实缺陷文件
    trueFaultFileLineList = trueFaultFile.readlines()
    excel = openpyxl.Workbook()  # 目标存储excel文件
    sheet = excel.worksheets[0]  # 表
    sheetHead = ['project_path', 'version', 'lines',
                 'statement', 'suspicious', 'faulty']
    for i in range(len(sheetHead)):  # 写入表头
        sheet.cell(1, i+1).value = sheetHead[i]
    excelLine = 2  # excel检索行
    for i in projectFileLineList:
        fileLineContentList = i.rstrip("\n").split('#')
        #去除$-start
        if '$' in fileLineContentList[1]:
            fileLineContentList[1] = fileLineContentList[1][:fileLineContentList[1].find('$')]
            sheet.cell(excelLine, 1).value = fileLineContentList[1]  # project_path
        else:
            sheet.cell(excelLine, 1).value = fileLineContentList[1]  # project_path
        #去除$-end
        sheet.cell(excelLine, 2).value = version  # version
        sheet.cell(excelLine, 3).value = fileLineContentList[2]  # lines
        # 获取代码-start
        lineOfCode = linecache.getline(
            CODE_FILE_PATH+fileLineContentList[1].replace(".", "/")+'.java', int(fileLineContentList[2]))
        # 获取代码-end
        sheet.cell(excelLine, 4).value = lineOfCode  # statement
        sheet.cell(excelLine, 5).value = fileLineContentList[3]  # suspicious
        # 判断该语句是否为真实缺陷-start
        isFault = False  # 是否为真实缺陷，默认为False
        for j in trueFaultFileLineList:
            trueFaultfileLineContentList = j.rstrip("\n").split('#')
            if fileLineContentList[1]+'.java' == trueFaultfileLineContentList[0] and fileLineContentList[2] == trueFaultfileLineContentList[1]:
                isFault = True
            else:
                isFault = False
        if isFault:
            sheet.cell(excelLine, 6).value = 1  # faulty
            print(project+'-'+str(version) +
                  fileLineContentList[1]+'-'+fileLineContentList[2]+' -> is true fault')
        else:
            sheet.cell(excelLine, 6).value = 0  # faulty
        # 判断该语句是否为真实缺陷-end
        excelLine += 1
    excel.save(TARGET_EXCEL_PATH+project+'-'+str(version)+'.xlsx')
    print(project+'-'+str(version)+'.xlsx'+' -> ok')
# 通过txt获取包含可疑度的excel-end


# 运行java文件生成json的AST树-start
def run_java_get_ASTjson(version, project):
    EclipseAST = '/Users/lvlaxjh/code/CBFL/AST/EclipseAST'
    AST_TO_JSON_APP_JAVA = '/Users/lvlaxjh/code/CBFL/AST/EclipseAST/App.java'
    EXCEL_PATH = '/Users/lvlaxjh/code/CBFL/data/%s/%s-%s.xlsx' % (
        project, project,str(version))#每个版本excel的路径
    JAVA_FATHER_PATH = '/Volumes/Elements/DSatr/%s-direct/%s-%s/source/' % (
        project, project, str(version))#java文件父路径
    excel = openpyxl.load_workbook(EXCEL_PATH)  # excel文件
    sheet = excel.worksheets[0]  # 表
    line = 2  # sheet 行
    javaFileList = []
    javaFileNames = []
    #代码文件路径-start
    while True:
        if sheet.cell(line, 1).value != None:
            filePath = str(sheet.cell(line, 1).value).replace(".", "/")
            if filePath not in javaFileList:
                javaFileNames.append(filePath.split('/')[-1])
                javaFileList.append(filePath)
        else:
            break
        line += 1
    #代码文件路径-end
    for i in range(len(javaFileList)):
        paths = ''
        for j in javaFileList[i].split('/')[:-1]:
            paths += j
            paths += '/'
        JAVA_FILE_PATH = JAVA_FATHER_PATH+paths
        JAVA_FILE = javaFileNames[i]
        file = open(AST_TO_JSON_APP_JAVA, 'r+')
        # 修改java代码中文件路径-start
        fileList = file.readlines()
        fileList[51] = '		'+'String MAIN_PATH = \"'+JAVA_FILE_PATH+'\";\n'
        fileList[52] = '		'+'String JAVA_FILE = \"'+JAVA_FILE+'\";\n'
        file.close()
        file = open(AST_TO_JSON_APP_JAVA, 'w+')
        file.writelines(fileList)
        file.close()
        # 修改java代码中文件路径-end
        osres = os.popen(  # 运行java
            'cd '+EclipseAST+' && javac -Xlint:deprecation -cp \'lib/*\' *.java && java -Xss4m -cp .:\'lib/*\' App "$@"')
        print(osres.read())
# 运行java文件生成json的AST树-end




if __name__ == "__main__":
    # 通过txt获取包含可疑度的excel-start
    for i in range(24, 27):#(1，项目总数+1)
    #     get_project_sus(i, 'chart')
        run_java_get_ASTjson(i, 'chart')
        # print('%s%s ---> ok'%('chart',str(i)))
        analytic_complexity.save_as_csv('/Users/lvlaxjh/code/CBFL/data/chart/chart-%s.xlsx'%(str(i)),
                        '/Users/lvlaxjh/code/CBFL/data/chart/csv/',
                        '/Volumes/Elements/DSatr/chart-direct/chart-%s/source/'%(str(i)),
                        i,'chart')
    # 通过txt获取包含可疑度的excel-end
    # ----------------------------------------------------------------

