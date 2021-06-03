'''
Author: your name
Date: 2021-05-28 08:43:20
LastEditTime: 2021-06-01 15:05:08
Description: 将不同版本的数据集合为excel(不去重)
'''

import openpyxl
PROJECT_NAME = 'time'
FATHER_PATH = '/Users/lvlaxjh/code/dataset/Data/excels_time/DStar/'
TARGET_EXCEL_PATH = '/Users/lvlaxjh/code/CBFL/data/time1/'
EXCEL_NAME = PROJECT_NAME+'.xlsx'
FILE_LIST_VID = [3,6,16,17]
contentList = []

for i in FILE_LIST_VID:
    excel = openpyxl.load_workbook(
        FATHER_PATH+'%s-%s/%s%s.xlsx' % (PROJECT_NAME, str(i), PROJECT_NAME, str(i)))  # 打开excel
    sheet = excel.worksheets[0]  # 表
    sheetLine = 2  # sheet 行
    while True:
        if sheet.cell(sheetLine, 1).value != None:
            contentList.append({'path': sheet.cell(sheetLine, 1).value,
                                'line': sheet.cell(sheetLine, 2).value,
                                'statement': sheet.cell(sheetLine, 3).value,
                                'dstar': sheet.cell(sheetLine, 4).value,
                                'faulty': sheet.cell(sheetLine, 5).value,
                                'vid': i})
            sheetLine += 1
        else:
            break
    excel.close()

print(len(contentList))
tarExcel = openpyxl.Workbook()
tarSheet = tarExcel.worksheets[0]  # 表
sheetHead = ['项目路径', '行数', 'statement', 'dstar', 'faulty', 'pid']
for i in range(len(sheetHead)):
    tarSheet.cell(1, i+1).value = sheetHead[i]
sheetLine = 2  # sheet 行
for i in contentList:
    tarSheet.cell(sheetLine, 1).value = i['path']
    tarSheet.cell(sheetLine, 2).value = i['line']
    tarSheet.cell(sheetLine, 3).value = i['statement']
    tarSheet.cell(sheetLine, 4).value = i['dstar']
    tarSheet.cell(sheetLine, 5).value = i['faulty']
    tarSheet.cell(sheetLine, 6).value = i['vid']
    sheetLine += 1
tarExcel.save(TARGET_EXCEL_PATH+EXCEL_NAME)
# print(contentList)
