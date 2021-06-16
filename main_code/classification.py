'''
Author: jhc
Date: 2021-05-25 18:47:20
LastEditTime: 2021-06-15 17:25:29
Description: classification
 ██████╗██╗      █████╗ ███████╗███████╗██╗███████╗██╗ ██████╗ █████╗ ████████╗██╗ ██████╗ ███╗   ██╗
██╔════╝██║     ██╔══██╗██╔════╝██╔════╝██║██╔════╝██║██╔════╝██╔══██╗╚══██╔══╝██║██╔═══██╗████╗  ██║
██║     ██║     ███████║███████╗███████╗██║█████╗  ██║██║     ███████║   ██║   ██║██║   ██║██╔██╗ ██║
██║     ██║     ██╔══██║╚════██║╚════██║██║██╔══╝  ██║██║     ██╔══██║   ██║   ██║██║   ██║██║╚██╗██║
╚██████╗███████╗██║  ██║███████║███████║██║██║     ██║╚██████╗██║  ██║   ██║   ██║╚██████╔╝██║ ╚████║
 ╚═════╝╚══════╝╚═╝  ╚═╝╚══════╝╚══════╝╚═╝╚═╝     ╚═╝ ╚═════╝╚═╝  ╚═╝   ╚═╝   ╚═╝ ╚═════╝ ╚═╝  ╚═══╝
'''
# from numpy import result_type
import pandas as pd
from scipy.sparse import data
# from sklearn.model_selection import train_test_split  # 分割验证集和训练集
# from sklearn import tree  # 树
# from sklearn.model_selection import KFold  # k折交叉验证
# import graphviz  # 绘图
#
from sklearn.tree import DecisionTreeClassifier  # 决策树
from sklearn.naive_bayes import GaussianNB  # 贝叶斯
from sklearn.neighbors import KNeighborsClassifier  # k近邻
from sklearn.ensemble import RandomForestRegressor  # 随机森林
from sklearn.svm import SVC  # 支持向量机
#
# from collections import Counter
from imblearn.over_sampling import SMOTE
from joblib import dump, load
# import numpy as np
import csv
import json
import openpyxl
from copy import deepcopy

FATHER_PATH = '/Users/lvlaxjh/code/'


def draw_tree(treeModel, feature_names, class_names, savePath):
    dot_data = tree.export_graphviz(
        treeModel, out_file=None, feature_names=feature_names, class_names=class_names, filled=True, rounded=True, special_characters=True)
    graph = graphviz.Source(dot_data)
    graph.render(savePath)


# 决策树-start
def tree(filePercentage, project, id, tree_depth, trainOrModel, if_SMOTE=True):  # 训练决策树
    # 测试集路径
    TEST_PATH = f'{FATHER_PATH}CBFL/data/{project}/csv/traintest/{str(filePercentage)}-{project}-test-{id}.csv'
    # 训练集路径
    TRAIN_PATH = f'{FATHER_PATH}CBFL/data/{project}/csv/traintest/{str(filePercentage)}-{project}-train-{id}.csv'
    # 模型存储父路径
    joblib_SAVE_PATH = f'{FATHER_PATH}CBFL/data/{project}/csv/traintest/model/'
    # 测试集-start
    test_data = pd.read_csv(TEST_PATH)
    test_x = test_data.iloc[:, [6, 7, 8, 9,
                                10, 11, 12, 13, 14, 15, 17]]  # 取测试数据
    test_y = test_data['accuracy']  # 取样本类标签
    # 测试集-end
    # 训练集-start
    train_data = pd.read_csv(TRAIN_PATH)
    train_x = train_data.iloc[:, [6, 7, 8, 9,
                                  10, 11, 12, 13, 14, 15, 17]]  # 取训练数据
    train_y = train_data['accuracy']  # 取样本类标签
    # 训练集-end
    if if_SMOTE:
        smo = SMOTE(random_state=1)  # MOTE
        train_x, train_y = smo.fit_resample(train_x, train_y)  # 平衡训练数据集
    if trainOrModel == 'train':
        model = DecisionTreeClassifier()  # 决策树模型初始化
    elif trainOrModel == 'model':
        model = load(joblib_SAVE_PATH +
                     f'{str(filePercentage)}tree{str(id)}.model')
    model.fit(train_x, train_y)  # 拟合模型
    dump(model, joblib_SAVE_PATH +
         f'{str(filePercentage)}tree{str(id)}.model')  # 存储模型
    predictRes = model.predict(test_x)  # 预测结果，返回列表
    predictRes = list(predictRes)  # numpy.array 类型转 list
    return predictRes  # list
# 决策树-end


# 贝叶斯-start
def bayes(filePercentage, project, id, trainOrModel, if_SMOTE=True):  # 训练决策树
    # 测试集路径
    TEST_PATH = f'{FATHER_PATH}CBFL/data/{project}/csv/traintest/{str(filePercentage)}-{project}-test-{id}.csv'
    # 训练集路径
    TRAIN_PATH = f'{FATHER_PATH}CBFL/data/{project}/csv/traintest/{str(filePercentage)}-{project}-train-{id}.csv'
    # 模型存储父路径
    joblib_SAVE_PATH = f'{FATHER_PATH}CBFL/data/{project}/csv/traintest/model/'
    # 测试集-start
    test_data = pd.read_csv(TEST_PATH)
    test_x = test_data.iloc[:, [6, 7, 8, 9,
                                10, 11, 12, 13, 14, 15, 17]]  # 取测试数据
    test_y = test_data['accuracy']  # 取样本类标签
    # 测试集-end
    # 训练集-start
    train_data = pd.read_csv(TRAIN_PATH)
    train_x = train_data.iloc[:, [6, 7, 8, 9,
                                  10, 11, 12, 13, 14, 15, 17]]  # 取训练数据
    train_y = train_data['accuracy']  # 取样本类标签
    # 训练集-end
    if if_SMOTE:
        smo = SMOTE(random_state=1)  # MOTE
        train_x, train_y = smo.fit_resample(train_x, train_y)  # 平衡训练数据集
    if trainOrModel == 'train':
        model = GaussianNB()  # 贝叶斯模型初始化
    elif trainOrModel == 'model':
        model = load(joblib_SAVE_PATH +
                     f'{str(filePercentage)}bayes{str(id)}.model')
    model.fit(train_x, train_y)  # 拟合模型
    dump(model, joblib_SAVE_PATH +
         f'{str(filePercentage)}bayes{str(id)}.model')  # 存储模型
    predictRes = model.predict(test_x)  # 预测结果，返回列表
    predictRes = list(predictRes)  # numpy.array 类型转 list
    return predictRes  # list
# 贝叶斯-end


# KNN-start
def KNN(filePercentage, project, id, trainOrModel, if_SMOTE=True):  # 训练决策树
    # 测试集路径
    TEST_PATH = f'{FATHER_PATH}CBFL/data/{project}/csv/traintest/{str(filePercentage)}-{project}-test-{id}.csv'
    # 训练集路径
    TRAIN_PATH = f'{FATHER_PATH}CBFL/data/{project}/csv/traintest/{str(filePercentage)}-{project}-train-{id}.csv'
    # 模型存储父路径
    joblib_SAVE_PATH = f'{FATHER_PATH}CBFL/data/{project}/csv/traintest/model/'
    # 测试集-start
    test_data = pd.read_csv(TEST_PATH)
    test_x = test_data.iloc[:, [6, 7, 8, 9,
                                10, 11, 12, 13, 14, 15, 17]]  # 取测试数据
    test_y = test_data['accuracy']  # 取样本类标签
    # 测试集-end
    # 训练集-start
    train_data = pd.read_csv(TRAIN_PATH)
    train_x = train_data.iloc[:, [6, 7, 8, 9,
                                  10, 11, 12, 13, 14, 15, 17]]  # 取训练数据
    train_y = train_data['accuracy']  # 取样本类标签
    # 训练集-end
    if if_SMOTE:
        smo = SMOTE(random_state=1)  # MOTE
        train_x, train_y = smo.fit_resample(train_x, train_y)  # 平衡训练数据集
    if trainOrModel == 'train':
        model = KNeighborsClassifier()  # k近邻模型初始化
    elif trainOrModel == 'model':
        model = load(joblib_SAVE_PATH +
                     f'{str(filePercentage)}knn{str(id)}.model')
    model.fit(train_x, train_y)  # 拟合模型
    dump(model, joblib_SAVE_PATH +
         f'{str(filePercentage)}knn{str(id)}.model')  # 存储模型
    predictRes = model.predict(test_x)  # 预测结果，返回列表
    predictRes = list(predictRes)  # numpy.array 类型转 list
    return predictRes  # list
# KNN-end


# 随机森林-start
def rsandomForest(filePercentage, project, id, trainOrModel, if_SMOTE=True):  # 训练决策树
    # 测试集路径
    TEST_PATH = f'{FATHER_PATH}CBFL/data/{project}/csv/traintest/{str(filePercentage)}-{project}-test-{id}.csv'
    # 训练集路径
    TRAIN_PATH = f'{FATHER_PATH}CBFL/data/{project}/csv/traintest/{str(filePercentage)}-{project}-train-{id}.csv'
    # 模型存储父路径
    joblib_SAVE_PATH = f'{FATHER_PATH}CBFL/data/{project}/csv/traintest/model/'
    # 测试集-start
    test_data = pd.read_csv(TEST_PATH)
    test_x = test_data.iloc[:, [6, 7, 8, 9,
                                10, 11, 12, 13, 14, 15, 17]]  # 取测试数据
    test_y = test_data['accuracy']  # 取样本类标签
    # 测试集-end
    # 训练集-start
    train_data = pd.read_csv(TRAIN_PATH)
    train_x = train_data.iloc[:, [6, 7, 8, 9,
                                  10, 11, 12, 13, 14, 15, 17]]  # 取训练数据
    train_y = train_data['accuracy']  # 取样本类标签
    # 训练集-end
    if if_SMOTE:
        smo = SMOTE(random_state=1)  # MOTE
        train_x, train_y = smo.fit_resample(train_x, train_y)  # 平衡训练数据集
    if trainOrModel == 'train':
        model = RandomForestRegressor()  # 随机森林模型初始化
    elif trainOrModel == 'model':
        model = load(joblib_SAVE_PATH +
                     f'{str(filePercentage)}forest{str(id)}.model')
    model.fit(train_x, train_y)  # 拟合模型
    dump(model, joblib_SAVE_PATH +
         f'{str(filePercentage)}forest{str(id)}.model')  # 存储模型
    resScore = model.score(test_x, test_y)  # 预测评分
    predictRes = model.predict(test_x)  # 预测结果，返回列表
    predictRes = list(predictRes)  # numpy.array 类型转 list
    return predictRes  # list
# 随机森林-end


# 支持向量机-start
def SVc(filePercentage, project, id, trainOrModel, if_SMOTE=True):  # 训练决策树
    # 测试集路径
    TEST_PATH = f'{FATHER_PATH}CBFL/data/{project}/csv/traintest/{str(filePercentage)}-{project}-test-{id}.csv'
    # 训练集路径
    TRAIN_PATH = f'{FATHER_PATH}CBFL/data/{project}/csv/traintest/{str(filePercentage)}-{project}-train-{id}.csv'
    # 模型存储父路径
    joblib_SAVE_PATH = f'{FATHER_PATH}CBFL/data/{project}/csv/traintest/model/'
    # 测试集-start
    test_data = pd.read_csv(TEST_PATH)
    test_x = test_data.iloc[:, [6, 7, 8, 9,
                                10, 11, 12, 13, 14, 15, 17]]  # 取测试数据
    test_y = test_data['accuracy']  # 取样本类标签
    # 测试集-end
    # 训练集-start
    train_data = pd.read_csv(TRAIN_PATH)
    train_x = train_data.iloc[:, [6, 7, 8, 9,
                                  10, 11, 12, 13, 14, 15, 17]]  # 取训练数据
    train_y = train_data['accuracy']  # 取样本类标签
    # 训练集-end
    if if_SMOTE:
        smo = SMOTE(random_state=1)  # MOTE
        train_x, train_y = smo.fit_resample(train_x, train_y)  # 平衡训练数据集
    if trainOrModel == 'train':
        model = SVC(kernel='poly')  # SVC模型初始化
    elif trainOrModel == 'model':
        model = load(joblib_SAVE_PATH +
                     f'{str(filePercentage)}svc{str(id)}.model')
    model.fit(train_x, train_y)  # 拟合模型
    dump(model, joblib_SAVE_PATH +
         f'{str(filePercentage)}svc{str(id)}.model')  # 存储模型
    predictRes = model.predict(test_x)  # 预测结果，返回列表
    predictRes = list(predictRes)  # numpy.array 类型转 list
    return predictRes  # list
# 支持向量机-end


# 将预测结果写入数据集 -start
def write_res_in_csv(project, id, resList):
    CSV_PATH = f'{FATHER_PATH}CBFL/data/{project}/csv/traintest/{project}-test-{str(id)}.csv'
    csvLine = 0  # csv行
    csvHead = []  # 存储csv文件头
    saveResList = []  # 存储所有文件
    with open(CSV_PATH) as f:
        csvFile = csv.reader(f)
        for row in csvFile:
            if csvLine == 0:
                csvHead = row
            else:
                row[20] = resList[csvLine-1]
                saveResList.append(row)
            csvLine += 1
    csvFile = open(CSV_PATH, 'w', encoding="utf-8", newline="")
    csvWriter = csv.writer(csvFile)
    csvWriter.writerow(csvHead)
    for i in saveResList:
        csvWriter.writerow(i)
    csvFile.close()
# 将预测结果写入数据集 -end


# 取得真实结果-start
def get_true_results(project, version, percentage):
    CSV_PATH = f'{FATHER_PATH}CBFL/data/{project}/csv/traintest/{percentage}-{project}-test-{str(version)}.csv'
    csvLine = 0  # csv行
    resList = []
    with open(CSV_PATH) as f:
        csvFile = csv.reader(f)
        for row in csvFile:
            if csvLine != 0:
                resList.append(int(row[19]))
            csvLine += 1
    return resList
# 取得真实结果-end


# 计算tp，tn，fp，fn-start
def get_tp_tn_fp_fn(project, id, predictRes, trueRes):
    # TP-start
    TP = 0
    TN = 0
    FP = 0
    FN = 0
    for i in range(len(predictRes)):
        if predictRes[i] == trueRes[i] == 1:
            TP += 1
        if predictRes[i] == trueRes[i] == 0:
            TN += 1
        if predictRes[i] == 1 and trueRes[i] == 0:
            FP += 1
        if predictRes[i] == 0 and trueRes[i] == 1:
            FN += 1
    # precesion = 0
    if TP+FP == 0:
        precesion = 0
    else:
        precesion = TP/(TP+FP)
    if TP+FN == 0:
        recall = 0
    else:
        recall = TP/(TP+FN)
    # recall = 0
    # print(f'{project}{id} TP : {TP} TN : {TN} FP : {FP} FN : {FN} precesion : {precesion} recall : {recall}')
    return TP, TN, FP, FN, precesion, recall
    # TP-end
# 计算tp，tn，fp，fn-end


# 保存结果-start
def save_result_to_excel(project, resList):
    excel = openpyxl.Workbook()  # 目标存储excel文件
    sheet = excel.worksheets[0]  # 表
    sheetHead = ['project', 'classification',
                 'SMOTE', 'filePercentage', 'TP', 'TN', 'FP', 'FN', 'pre', 'recall', 'f1']
    for i in range(len(sheetHead)):  # 写入表头
        sheet.cell(1, i+1).value = sheetHead[i]
    excelLine = 2  # excel检索行
    for i in resList:
        sheet.cell(excelLine, 1).value = i['project']
        sheet.cell(excelLine, 2).value = i['classification']
        sheet.cell(excelLine, 3).value = i['SMOTE']
        sheet.cell(excelLine, 4).value = i['filePercentage']
        sheet.cell(excelLine, 5).value = i['TP']
        sheet.cell(excelLine, 6).value = i['TN']
        sheet.cell(excelLine, 7).value = i['FP']
        sheet.cell(excelLine, 8).value = i['FN']
        sheet.cell(excelLine, 9).value = i['pre']
        sheet.cell(excelLine, 10).value = i['recall']
        sheet.cell(excelLine, 11).value = i['f1']
        #
        excelLine += 1
    excel.save(f'{FATHER_PATH}CBFL/data/{project}'+'.xlsx')
# 保存结果-end


def save_predictRes(project, version, percentage, predictRes):
    CSV_PATH = f'{FATHER_PATH}CBFL/data/{project}/csv/traintest/{percentage}-{project}-test-{str(version)}.csv'
    csvLine = 0  # csv行
    resList = []
    #读原csv-start
    with open(CSV_PATH) as f:
        csvFile = csv.reader(f)
        for row in csvFile:
            resList.append(row)
            # csvLine += 1
    csvLine = 0  # csv行
    #读原csv-end
    csvFile = open(CSV_PATH, 'w', encoding="utf-8", newline="")
    csvWriter = csv.writer(csvFile)
    for i in resList:
        if csvLine ==0:
            csvWriter.writerow(i)
        else:
            data_TEM = deepcopy(i)
            data_TEM[20] = predictRes[csvLine-1]
            csvWriter.writerow(data_TEM)
        csvLine+=1
    csvFile.close()



if __name__ == "__main__":
    # for project in ['chart', 'lang', 'time', 'math', 'closure', 'mockito']:
    for project in ['mockito']:
        # project = 'chart'  # *
        settingJson = open(f'{FATHER_PATH}CBFL/main_code/setting.json', 'r')
        settingContent = settingJson.read()
        setting = json.loads(settingContent)
        k = setting[project]['k']  # k折交叉
        percentageList = setting['percentageList']  # 前百分比数据列表
        useModel = 'train'  # train/model ***
        #
        totalTP = 0
        totalTn = 0
        totalFP = 0
        totalFN = 0
        totalprecesion = 0
        totalrecall = 0
        #
        resList = []
        resDict = {
            'project': project,
            'classification': '',
            'SMOTE': '',
            'filePercentage': '',
            'TP': '',
            'TN': '',
            'FP': '',
            'FN': '',
            'pre': '',
            'recall': '',
            'f1': '',
        }
        #
        print(f'\033[1;35m----- >{project}< ----- \033[0m')
        # for func in ['tree', 'bayes', 'KNN', 'rsandomForest', 'SVc']:
        for func in ['tree', 'bayes', 'KNN', 'rsandomForest', 'SVc']:
            # for func in ['bayes']:
            print(f'\033[1;34m----- >{func}< ----- \033[0m')
            resDict['classification'] = deepcopy(func)
            for smote in [True, False]:
                print(f'\033[1;35mSMOTE : {str(smote)} \033[0m')
                resDict['SMOTE'] = deepcopy(str(smote))
                for filePercentage in percentageList:
                    print(
                        f'\033[1;35mfilePercentage : {str(filePercentage)} \033[0m')
                    resDict['filePercentage'] = deepcopy(filePercentage)
                    for i in range(k):
                        if func == 'tree':
                            predictRes = tree(
                                filePercentage, project, i, tree_depth=3, trainOrModel=useModel, if_SMOTE=smote)  # 决策树
                        if func == 'bayes':
                            predictRes = bayes(
                                filePercentage, project, i, trainOrModel=useModel, if_SMOTE=smote)  # 贝叶斯
                            save_predictRes(
                                project, i, filePercentage, predictRes)
                        if func == 'KNN':
                            predictRes = KNN(filePercentage, project, i,
                                             trainOrModel=useModel, if_SMOTE=smote)  # K近邻
                        if func == 'rsandomForest':
                            predictRes = rsandomForest(filePercentage, project, i, trainOrModel=useModel,
                                                       if_SMOTE=smote)  # 随机森林
                        if func == 'SVc':
                            predictRes = SVc(filePercentage, project, i, trainOrModel=useModel,
                                             if_SMOTE=smote)  # 支持向量机
                        trueRes = get_true_results(
                            project, i, filePercentage)  # 真实数据
                        TP, TN, FP, FN, precesion, recall = get_tp_tn_fp_fn(
                            project, i, predictRes, trueRes)
                        totalTP += TP
                        totalTn += TN
                        totalFP += FP
                        totalFN += FN
                    if totalTP+totalFP == 0:
                        totalprecesion = 0
                    else:
                        totalprecesion = totalTP/(totalTP+totalFP)
                    if totalTP+totalFN == 0:
                        totalrecall = 0
                    else:
                        totalrecall = recall = totalTP/(totalTP+totalFN)
                    if 2*totalTP+totalFP+totalFN == 0:
                        F1 = 0
                    else:
                        F1 = 2*totalTP/(2*totalTP+totalFP+totalFN)
                    print(
                        f'{project}total TP : {totalTP} TN : {totalTn} FP : {totalFP} FN : {totalFN} precesion : {totalprecesion} recall : {totalrecall} F1 : {F1}')
                    resDict['TP'] = deepcopy(totalTP)
                    resDict['TN'] = deepcopy(totalTn)
                    resDict['FP'] = deepcopy(totalFP)
                    resDict['FN'] = deepcopy(totalFN)
                    resDict['pre'] = deepcopy(totalprecesion)
                    resDict['recall'] = deepcopy(totalrecall)
                    resDict['f1'] = deepcopy(F1)
                    resList.append(deepcopy(resDict))
                    #
                    totalTP = 0
                    totalTn = 0
                    totalFP = 0
                    totalFN = 0
                    totalprecesion = 0
                    totalrecall = 0
                    #
        save_result_to_excel(project, resList)
