'''
Author: jhc
Date: 2021-06-15 16:04:17
LastEditTime: 2021-07-13 15:37:12
Description: 分类器分类用于并列可疑度数据
'''
"""
 ██████╗██╗      █████╗ ███████╗███████╗██╗███████╗██╗ ██████╗ █████╗ ████████╗██╗ ██████╗ ███╗   ██╗
██╔════╝██║     ██╔══██╗██╔════╝██╔════╝██║██╔════╝██║██╔════╝██╔══██╗╚══██╔══╝██║██╔═══██╗████╗  ██║
██║     ██║     ███████║███████╗███████╗██║█████╗  ██║██║     ███████║   ██║   ██║██║   ██║██╔██╗ ██║
██║     ██║     ██╔══██║╚════██║╚════██║██║██╔══╝  ██║██║     ██╔══██║   ██║   ██║██║   ██║██║╚██╗██║
╚██████╗███████╗██║  ██║███████║███████║██║██║     ██║╚██████╗██║  ██║   ██║   ██║╚██████╔╝██║ ╚████║
 ╚═════╝╚══════╝╚═╝  ╚═╝╚══════╝╚══════╝╚═╝╚═╝     ╚═╝ ╚═════╝╚═╝  ╚═╝   ╚═╝   ╚═╝ ╚═════╝ ╚═╝  ╚═══╝
████████╗██╗███████╗██████╗
╚══██╔══╝██║██╔════╝██╔══██╗
   ██║   ██║█████╗  ██║  ██║
   ██║   ██║██╔══╝  ██║  ██║
   ██║   ██║███████╗██████╔╝
   ╚═╝   ╚═╝╚══════╝╚═════╝
"""


# from numpy import result_type

#
# from collections import Counter

# from sklearn.model_selection import train_test_split  # 分割验证集和训练集
# from sklearn import tree  # 树
# from sklearn.model_selection import KFold  # k折交叉验证
# import graphviz  # 绘图
#

# from scipy.sparse import data


# from joblib import dump, load


import csv
import json
from copy import deepcopy
import openpyxl
import pandas as pd
from imblearn.over_sampling import SMOTE  # smote
from imblearn.over_sampling import RandomOverSampler  # 随机过采样
from imblearn.under_sampling import RandomUnderSampler  # 随机欠采样
from sklearn.ensemble import RandomForestClassifier  # 随机森林
from sklearn.naive_bayes import GaussianNB  # 贝叶斯
from sklearn.neighbors import KNeighborsClassifier  # k近邻
from sklearn.svm import SVC  # 支持向量机
from sklearn.tree import DecisionTreeClassifier  # 决策树
FATHER_PATH = "C:/ssdcode/"


def use_classification(TEST_PATH, TRAIN_PATH, classificationFunc, k_Id, tiedK, sampling):
    # 测试集-start
    test_data = pd.read_csv(TEST_PATH)
    test_x = test_data.iloc[:, [6, 7, 8, 9, 10, 11, 12, 13, 14, 17]]  # 取测试数据
    # test_y = test_data["accuracy"]  # 取样本类标签
    # 测试集-end
    # 训练集-start
    train_data = pd.read_csv(TRAIN_PATH)
    train_x = train_data.iloc[:, [6, 7, 8, 9, 10, 11, 12, 13, 14, 17]]  # 取训练数据
    # train_y = train_data["accuracy"]  # 取样本类标签
    train_y = train_data.iloc[:, [19]]  # 取样本类标签
    # 训练集-end
    # 选择采样-start
    if sampling == "smote":
        sam = SMOTE()  # MOTE
        train_x, train_y = sam.fit_resample(train_x, train_y)  # 平衡训练数据集
    elif sampling == "over":
        sam = RandomOverSampler()  # 随机过采样
        train_x, train_y = sam.fit_resample(train_x, train_y)  # 平衡训练数据集
    elif sampling == "under":
        sam = RandomUnderSampler()  # 随机欠采样
        train_x, train_y = sam.fit_resample(train_x, train_y)  # 平衡训练数据集
    # 选择采样-end
    if classificationFunc == "tree":
        model = DecisionTreeClassifier()  # 决策树模型初始化
    elif classificationFunc == "bayes":
        model = GaussianNB()  # 贝叶斯模型初始化
    elif classificationFunc == "KNN":
        model = KNeighborsClassifier()  # K近邻模型初始化
    elif classificationFunc == "randomForest":
        model = RandomForestClassifier()  # 随机森林模型初始化
    elif classificationFunc == "SVC":
        model = SVC()  # SVC模型初始化
    model.fit(train_x, train_y.values.ravel())  # 拟合模型
    predictRes = model.predict(test_x)  # 预测结果，返回列表
    predictRes = list(predictRes)  # numpy.array 类型转 list
    return predictRes


# 取得真实结果-start
def get_true_results(TEST_PATH):
    resList = []
    with open(TEST_PATH) as f:
        csvFile = csv.reader(f)
        for row in csvFile:
            if row[0] != "dataset":
                resList.append(int(deepcopy(row[19])))
    return resList


# 计算tp，tn，fp，fn-start
def get_tp_tn_fp_fn(predictRes, trueRes):
    # TP-start
    TP = 0
    TN = 0
    FP = 0
    FN = 0
    for i in range(len(predictRes)):
        if predictRes[i] == trueRes[i] == 1:
            TP += 1
        if predictRes[i] == trueRes[i] == 0:
            TN += 1
        if predictRes[i] == 1 and trueRes[i] == 0:
            FP += 1
        if predictRes[i] == 0 and trueRes[i] == 1:
            FN += 1
    # precesion = 0
    if TP + FP == 0:
        precesion = 0
    else:
        precesion = TP / (TP + FP)
    if TP + FN == 0:
        recall = 0
    else:
        recall = TP / (TP + FN)
    # recall = 0
    # print(f'{project}{id} TP : {TP} TN : {TN} FP : {FP} FN : {FN} precesion : {precesion} recall : {recall}')
    return TP, TN, FP, FN, precesion, recall
    # TP-end


# 计算tp，tn，fp，fn-end


# 保存结果-start
def save_result_to_excel(project, resList):
    excel = openpyxl.Workbook()  # 目标存储excel文件
    sheet = excel.worksheets[0]  # 表
    sheetHead = ["project", "classification", "sam", "tiedK", "k", "TP", "TN", "FP", "FN", "pre", "recall", "f1"]
    for i in range(len(sheetHead)):  # 写入表头
        sheet.cell(1, i + 1).value = sheetHead[i]
    excelLine = 2  # excel检索行
    for i in resList:
        sheet.cell(excelLine, 1).value = i["project"]
        sheet.cell(excelLine, 2).value = i["classification"]
        sheet.cell(excelLine, 3).value = i["sam"]
        sheet.cell(excelLine, 4).value = i["tiedK"]
        sheet.cell(excelLine, 5).value = i["k"]
        sheet.cell(excelLine, 6).value = i["TP"]
        sheet.cell(excelLine, 7).value = i["TN"]
        sheet.cell(excelLine, 8).value = i["FP"]
        sheet.cell(excelLine, 9).value = i["FN"]
        sheet.cell(excelLine, 10).value = i["pre"]
        sheet.cell(excelLine, 11).value = i["recall"]
        sheet.cell(excelLine, 12).value = i["f1"]
        #
        excelLine += 1
    excel.save(f"{FATHER_PATH}CBFL/dataset/result/tiedK_{project}" + ".xlsx")


# 保存结果-end


def save_predictRes(project, version, percentage, predictRes):
    CSV_PATH = f"{FATHER_PATH}CBFL/data/{project}/csv/traintest/{percentage}-{project}-test-{str(version)}.csv"
    csvLine = 0  # csv行
    resList = []
    # 读原csv-start
    with open(CSV_PATH) as f:
        csvFile = csv.reader(f)
        for row in csvFile:
            resList.append(row)
            # csvLine += 1
    csvLine = 0  # csv行
    # 读原csv-end
    csvFile = open(CSV_PATH, "w", encoding="utf-8", newline="")
    csvWriter = csv.writer(csvFile)
    for i in resList:
        if csvLine == 0:
            csvWriter.writerow(i)
        else:
            data_TEM = deepcopy(i)
            data_TEM[20] = predictRes[csvLine - 1]
            csvWriter.writerow(data_TEM)
        csvLine += 1
    csvFile.close()


def save_result_in_one_excel():
    projectList = ['chart', 'lang', 'time', 'math', 'mockito']
    # projectList = ['chart']
    allResult = []
    for project in projectList:
        excel = openpyxl.load_workbook(f'{FATHER_PATH}CBFL/dataset/result/tiedK_{project}.xlsx')
        sheet = excel.worksheets[0]  # 表
        excelLine = 2
        while True:
            if sheet.cell(excelLine, 1).value is not None:
                list_TEM = []
                for col in range(1, 13):
                    list_TEM.append(deepcopy(sheet.cell(excelLine, col).value))
                allResult.append(deepcopy(list_TEM))
                excelLine+=1
            else:
                break
        excel.close()
    excel = openpyxl.Workbook()  # 目标存储excel文件
    sheet = excel.worksheets[0]
    sheetHead = ["project", "classification", "sam", "tiedK", "k", "TP", "TN", "FP", "FN", "pre", "recall", "f1"]
    for i in range(len(sheetHead)):  # 写入表头
        sheet.cell(1, i + 1).value = sheetHead[i]
    excelLine = 2
    for i in allResult:
        for j in range(1, 13):
            sheet.cell(excelLine, j).value = i[j-1]
        #
        excelLine += 1
    excel.save(f"{FATHER_PATH}CBFL/dataset/result/tiedK" + ".xlsx")


if __name__ == "__main__":
    # for project in ['chart', 'lang', 'time', 'math','mockito', 'closure' ]:
    # for project in ['chart']:
    for project in ['chart', 'lang', 'time', 'math', 'mockito']:
        # for project in ["mockito"]:
        settingJson = open(f"{FATHER_PATH}CBFL/code/setting.json", "r", encoding="utf-8")
        settingContent = settingJson.read()
        setting = json.loads(settingContent)
        KList = setting["k"]  # k折交叉
        percentageList = setting["percentageList"]  # 前百分比数据列表
        tiedKList = setting["tiedK"]
        # useModel = "train"  # train/model ***
        #
        totalTP = 0
        totalTn = 0
        totalFP = 0
        totalFN = 0
        totalprecesion = 0
        totalrecall = 0
        #
        resList = []
        resDict = {
            "project": project,
            "classification": "",
            "SMOTE": "",
            "tiedK": "",
            "k": "",
            "TP": "",
            "TN": "",
            "FP": "",
            "FN": "",
            "pre": "",
            "recall": "",
            "f1": "",
        }
        #
        # for classificationFunc in ["tree", "bayes", "KNN", "randomForest", "SVC"]:
        for classificationFunc in ["bayes"]:
            resDict["classification"] = deepcopy(classificationFunc)
            for sampling in ["none", "smote", "over", "under"]:
                resDict["sam"] = deepcopy(sampling)
                for tiedK in tiedKList:
                    resDict["tiedK"] = deepcopy(tiedK)
                    for k in KList:
                        resDict["k"] = deepcopy(k)
                        for k_Id in range(k):
                            TEST_PATH = f"{FATHER_PATH}CBFL/dataset/{project}/tiedK_traintest/test-{project}-{tiedK}-{k}-{k_Id}.csv"
                            TRAIN_PATH = f"{FATHER_PATH}CBFL/dataset/{project}/tiedK_traintest/train-{project}-{tiedK}-{k}-{k_Id}.csv"
                            predictRes = use_classification(TEST_PATH, TRAIN_PATH, classificationFunc, k_Id, tiedK, sampling)
                            trueRes = get_true_results(TEST_PATH)  # 获取真实数据以计算F1等
                            TP, TN, FP, FN, precesion, recall = get_tp_tn_fp_fn(predictRes, trueRes)
                            totalTP += TP
                            totalTn += TN
                            totalFP += FP
                            totalFN += FN
                        if totalTP + totalFP == 0:
                            totalprecesion = 0
                        else:
                            totalprecesion = totalTP / (totalTP + totalFP)
                        if totalTP + totalFN == 0:
                            totalrecall = 0
                        else:
                            totalrecall = totalTP / (totalTP + totalFN)
                        if 2 * totalTP + totalFP + totalFN == 0:
                            F1 = 0
                        else:
                            F1 = 2 * totalTP / (2 * totalTP + totalFP + totalFN)
                        print(f"\033[1;35mclassificationFunc : {classificationFunc} sampling : {sampling} tiedK : {tiedK} k : {k} \033[0m")
                        print(f"{project} total TP : {totalTP} TN : {totalTn} FP : {totalFP} FN : {totalFN} precesion : {totalprecesion} recall : {totalrecall} F1 : {F1}")
                        resDict["TP"] = deepcopy(totalTP)
                        resDict["TN"] = deepcopy(totalTn)
                        resDict["FP"] = deepcopy(totalFP)
                        resDict["FN"] = deepcopy(totalFN)
                        resDict["pre"] = deepcopy(totalprecesion)
                        resDict["recall"] = deepcopy(totalrecall)
                        resDict["f1"] = deepcopy(F1)
                        resList.append(deepcopy(resDict))
                        #
                        totalTP = 0
                        totalTn = 0
                        totalFP = 0
                        totalFN = 0
                        totalprecesion = 0
                        totalrecall = 0
                        #
        save_result_to_excel(project, resList)
    save_result_in_one_excel()
